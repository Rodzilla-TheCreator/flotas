from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, json, send_file, current_app
from flask_mail import Mail, Message
import mysql.connector
from datetime import datetime, timedelta, time
from hdbcli import dbapi as hana
import logging
from logging.handlers import RotatingFileHandler, SMTPHandler
import traceback
import pandas as pd
from urllib.parse import urlparse, urlunparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os
import shutil
import gc
from threading import Timer
import re
import win32com.client
import smtplib
from apscheduler.schedulers.background import BackgroundScheduler
import imaplib
import email
from email.header import decode_header
import time


logging.basicConfig(level=logging.DEBUG)


app = Flask(__name__)
app.secret_key = 'carepa'




# Configure logging
if not app.debug:
    file_handler = RotatingFileHandler('error.log', maxBytes=1024 * 1024 * 100, backupCount=20)
    file_handler.setLevel(logging.ERROR)
    formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]')
    file_handler.setFormatter(formatter)
    app.logger.addHandler(file_handler)



@app.before_request
def log_request_info():
    url_parts = list(urlparse(request.url))
    # Combine the parts to get the full URL
    full_url = urlunparse(url_parts)
    if not (request.path.endswith('favicon.ico') or request.path.endswith('.css') or request.path.endswith('.png') or '/get_order_data/' in request.path or '/get_active_mechanics/' in request.path or '/get_assigned_mechanics/' in request.path):
        logging.info(f'Route accessed: {full_url}, Method: {request.method}, Time: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, IP: {request.remote_addr}')

@app.errorhandler(Exception)
def handle_exception(e):
    tb = traceback.format_exc()
    app.logger.error('Unhandled Exception: %s\n%s', e, tb)
    return "Internal Server Error", 500

@app.errorhandler(404)
def page_not_found(e):
    if request.path != '/favicon.ico':
        app.logger.error('Page not found: %s\n%s', request.url, str(e))
    return "404 Not Found: The requested URL was not found on the server.", 404





#to download dependencies on wifi: pip install Flask==3.0.2 mysql-connector-python==8.0.29 Werkzeug==3.0.1 Jinja2==3.1.3 itsdangerous==2.1.2 click==8.1.7 MarkupSafe==2.1.5 hdbcli==2.19.21 waitress==3.0.0

# to download dependencies without wifi: 
# pip install --no-index --find-links="C:\Users\rodrigo.monterroso\Downloads\flotas first running 7.9.2024\flotas Upload 30 6 24\Flotas" numpy-2.0.0-cp312-cp312-win_amd64.whl pandas-2.2.2-cp312-cp312-win_amd64.whl python_dateutil-2.9.0.post0-py2.py3-none-any.whl pytz-2024.1-py2.py3-none-any.whl tzdata-2024.1-py2.py3-none-any.whl six-1.16.0-py2.py3-none-any.whl blinker-1.7.0-py3-none-any.whl click-8.1.7-py3-none-any.whl colorama-0.4.6-py2.py3-none-any.whl Flask-3.0.2-py3-none-any.whl hdbcli-2.19.21-cp36-cp36m-win_amd64.whl itsdangerous-2.1.2-py3-none-any.whl Jinja2-3.1.3-py3-none-any.whl MarkupSafe-2.1.5-cp312-cp312-win_amd64.whl mysql_connector_python-8.0.30-cp312-cp312-win_amd64.whl waitress-3.0.0-py3-none-any.whl Werkzeug-3.0.1-py3-none-any.whl openpyxl-3.1.5-py2.py3-none-any.whl et_xmlfile-1.1.0-py3-none-any.whl Flask_SocketIO-5.3.6-py3-none-any.whl python_socketio-5.11.3-py3-none-any.whl bidict-0.23.1-py3-none-any.whl python_engineio-4.9.1-py3-none-any.whl simple_websocket-1.0.0-py3-none-any.whl wsproto-1.2.0-py3-none-any.whl h11-0.14.0-py3-none-any.whl requests-2.32.3-py3-none-any.whl charset_normalizer-3.3.2-py3-none-any.whl idna-3.7-py3-none-any.whl urllib3-2.2.2-py3-none-any.whl certifi-2024.7.4-py3-none-any.whl pywin32-306-cp312-cp312-win_amd64.whl flask_mail-0.10.0-py3-none-any.whl APScheduler-3.10.4-py3-none-any.whl tzlocal-5.2-py3-none-any.whl
# pip install --no-index --find-links="C:\Users\rodrigo.monterroso\Downloads\flotas first running 7.9.2024\flotas Upload 30 6 24\Flotas" 


#to start page
#.\venv\Scripts\Activate
#flask run --debug
#
def get_db_connection():
    """Establish connection with the database."""
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='',
        database='MechanicOrganizationalSystem',
    )


def connect_hana(server='172.16.31.16', puerto=30015, usuario='ADMIN_FLOTA', clave ='Montasa2024') -> hana.Connection:
    """
    Función que se encarga de realizar la conexión con la base de datos en HANA.
    Recibe como parámetro la dirección IP, puerto, usuario y clave del servidor.

    Utiliza la librería hdbcli, módulo dbapi.
    """
    try:
        connection = hana.connect(address=server, port=puerto, user=usuario, password=clave)
        print("Conexión exitosa a HANA DB.")
        return connection
    except hana.Error as e:
        print(f"Error al conectar a HANA DB: {e}")
connection = connect_hana()






@app.route('/')
def home():
    return render_template('home.html')


# Configuración de la aplicación Flask y Flask-Mail
app.config['MAIL_SERVER'] = 'smtp.mail.yahoo.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'erroresmail@yahoo.com'
app.config['MAIL_PASSWORD'] = 'lkoepebsefjsivei'  # Contraseña de aplicación
app.config['MAIL_DEFAULT_SENDER'] = ('LogiFleet', 'erroresmail@yahoo.com') 

mail = Mail(app)


IMAP_SERVER = 'imap.mail.yahoo.com'
EMAIL_USER = 'erroresmail@yahoo.com'  # The same email you used in the configuration
EMAIL_PASS = 'lkoepebsefjsivei'

# Check email and process horometro updates or incident reports
def check_email():
    # Conectar al servidor IMAP y loguearse
    with imaplib.IMAP4_SSL(IMAP_SERVER) as mail_server:
        mail_server.login(EMAIL_USER, EMAIL_PASS)
        mail_server.select("inbox")  # Seleccionar la bandeja de entrada

        # Buscar correos no leídos con asunto "1" (horometro) o "inc" (incidencias)
        status_horometro, messages_horometro = mail_server.search(None, '(UNSEEN SUBJECT "1")')
        status_incident, messages_incident = mail_server.search(None, '(UNSEEN SUBJECT "inc")')

        # Combinar los mensajes encontrados
        email_ids = messages_horometro[0].split() + messages_incident[0].split()

        for email_id in email_ids:
            # Obtener el correo
            status, msg_data = mail_server.fetch(email_id, "(RFC822)")
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])

                    # Decodificar el asunto del correo
                    subject, encoding = decode_header(msg["Subject"])[0]
                    if isinstance(subject, bytes):
                        subject = subject.decode(encoding if encoding else 'utf-8')

                    sender = msg.get("From")
                    print(f"New email received from {sender}: {subject}")

                    # Procesar según el asunto
                    if subject == '1':
                        process_horometro_update(msg)
                    elif subject == 'inc':
                        process_incident_creation(msg)

                    # Marcar el correo como leído
                    mail_server.store(email_id, '+FLAGS', '\\Seen')


# Process horometro updates
def process_horometro_update(msg):
    # Extraer vehicle_name y new_horometro desde el cuerpo del email
    body = msg.get_payload(decode=True).decode('utf-8')
    lines = body.splitlines()
    vehicle_name = lines[0].split(": ")[1]
    new_horometro = float(lines[1].split(": ")[1])

    # Llamar a la función para actualizar el horómetro en la base de datos
    response = update_horometros_via_email(vehicle_name, new_horometro)
    print(response)


# Process incident creation
def process_incident_creation(msg):
    # Extract vehicle name and incident description from email body
    body = msg.get_payload(decode=True).decode('utf-8')
    lines = body.splitlines()
    vehicle_name = lines[0].split(": ")[1]
    incident_description = lines[1].split(": ")[1]

    # Call the create incident route using the extracted data
    response = create_incident(vehicle_name, incident_description)
    print(f"Processed incident creation: {response}")

# Dummy function for creating an incident
def create_incident(vehicle_name, description):
    # Add the logic for creating an incident in the Flotas App's database
    # Example of a successful incident creation
    return f"Incident for {vehicle_name} created with description: {description}"

@app.route('/check_emails', methods=['GET'])
def check_emails_route():
    check_email()
    return "Checked emails for 'Hello World'. Check console for details."

# Configuración del registro
logging.basicConfig(level=logging.DEBUG)


def send_horometro_update(vehicle_name, new_horometro):
    # Abrir un contexto de aplicación
    with app.app_context():
        subject = "11"  # Asunto para la actualización de horómetro
        body = f"Vehicle: {vehicle_name}\nHorometro: {new_horometro}"

        # Crear el mensaje y enviarlo
        msg = Message(subject=subject, body=body, recipients=['erroresmail@yahoo.com'])  # Email de LogiFleet
        try:
            logging.info(f"Sending horometro update email for {vehicle_name} with horometro {new_horometro}...")
            mail.send(msg)
            logging.info("Email sent successfully")
        except Exception as e:
            logging.error(f"Failed to send horometro update email: {e}")



def generate_vehicle_report(company):
    original_file_path = 'C:\\Users\\rodrigo.monterroso\\Downloads\\flotas first running 7.9.2024\\flotas Upload 30 6 24\\Flotas\\15-07-2024 REPORTE UBICACION MONTASA -.xlsx'
    copied_file_path = 'C:\\Users\\rodrigo.monterroso\\Downloads\\flotas first running 7.9.2024\\flotas Upload 30 6 24\\Flotas\\temp_REPORTE_UBICACION_MONTASA.xlsx'
    
    shutil.copyfile(original_file_path, copied_file_path)

    conn = get_db_connection()
    cursor = conn.cursor()

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    cursor.execute(f'''
        SELECT VehicleName, Observacion, Status, Ubicacion
        FROM {vehicle_table}
        WHERE VehicleName NOT LIKE 'C%' AND VehicleName != 'generico'
    ''')
    vehicles = cursor.fetchall()
    
    try:
        workbook = load_workbook(copied_file_path, data_only=True, read_only=True)
        sheet = workbook.active

        vehicle_row_mapping = {}
        for row in sheet.iter_rows(min_row=8, max_row=196, max_col=26):
            vehicle_name = row[7].value
            if vehicle_name:
                vehicle_row_mapping[vehicle_name] = row

        workbook.close()
        workbook = load_workbook(copied_file_path)
        sheet = workbook.active

        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for vehicle in vehicles:
            vehicle_name = vehicle[0]
            if vehicle_name in vehicle_row_mapping:
                row = vehicle_row_mapping[vehicle_name]
                try:
                    sheet.cell(row=row[0].row, column=14).value = '1' if vehicle[2] in ['En Taller', 'Deshabilitado'] else '2' if vehicle[2] in ['En Renta', 'Reparacion Externa'] else '3' if vehicle[2] == 'Disponible' else ''
                except Exception as e:
                    logging.error(f"Error updating ST for {vehicle_name}: {e}")
                try:
                    sheet.cell(row=row[0].row, column=15).value = vehicle[1]
                except Exception as e:
                    logging.error(f"Error updating OBSERVACIONES for {vehicle_name}: {e}")
                try:
                    sheet.cell(row=row[0].row, column=18).value = vehicle[3]
                except Exception as e:
                    logging.error(f"Error updating UBICACIÓN for {vehicle_name}: {e}")

                try:
                    if vehicle[2] == 'Deshabilitado':
                        for cell in row:
                            sheet.cell(row=row[0].row, column=cell.col_idx).fill = red_fill
                        sheet.cell(row=row[0].row, column=23).value = 1
                    elif 'VENDIDO' in vehicle[1]:
                        for cell in row:
                            sheet.cell(row=row[0].row, column=cell.col_idx).fill = yellow_fill
                except Exception as e:
                    logging.error(f"Error updating highlight for {vehicle_name}: {e}")

        gc.collect()
        current_date = datetime.now().strftime('%d-%m-%Y')
        new_file_path = os.path.join(os.path.dirname(copied_file_path), f'{current_date} REPORTE UBICACION MONTASA -.xlsx')
        workbook.save(new_file_path)
        cursor.close()
        conn.close()

        return new_file_path

    except Exception as e:
        logging.error(f"Error processing the request: {e}")
        return None

def send_daily_report():
    with app.app_context():
        logging.debug('Entered send_daily_report function')
        company = 'MontasaHN'  # Adjust this based on your needs
        file_path = generate_vehicle_report(company)
        if not file_path:
            logging.error('Failed to generate vehicle report')
            return

        try:
            recipients = [
                "ventas5@montasa.com", "ventas3@montasa.com", "ventas2@montasa.com", "ventas1@montasa.com",
                "Fabrizio.monterroso@montasa.com", "supervisor.inventarios@montasa.com", "supervisor.logistica@montasa.com",
                "rodrigo.j.monterroso@gmail.com", "jefe.taller@montasa.com", "omar@montasa.com", "compras3@montasa.com",
                "contador@montasa.com", "gabriel.monterroso@montasa.com", "monitoreomontasa1@segemsahn.com",
                "taller.choloma@montasa.com", "elia.lopez@montasa.com", "logistica1@montasa.com"
            ]

            current_date = datetime.now().strftime('%d-%m-%Y')
            msg = Message(
                subject=f"Reporte de ubicacion MontasaHN {current_date}",
                body="",
                recipients=recipients
            )

            with app.open_resource(file_path) as fp:
                msg.attach(filename=os.path.basename(file_path), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", data=fp.read())

            logging.debug('Sending email with attachment')
            mail.send(msg)
            logging.debug('Email sent successfully')
        except Exception as e:
            logging.error(f"Error occurred: {e}")

# Set up the scheduler
scheduler = BackgroundScheduler()
scheduler.add_job(func=send_daily_report, trigger='cron', hour=17, minute=10)  # Schedule for 5:10 PM
#scheduler.add_job(func=check_email, trigger='interval', seconds=30) 
def initialize_scheduler():
    if not scheduler.running:
        scheduler.start()
        logging.info("Scheduler started")





        
@app.route('/get_assigned_mechanics/<order_id>', methods=['GET'])
def get_assigned_mechanics(order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch mechanics who have been assigned to the order, regardless of EndTime
    cursor.execute('''
        SELECT DISTINCT m.Name
        FROM Mechanics m
        JOIN TimeTracking t ON m.MechanicID = t.MechanicID
        WHERE t.OrderID = %s
    ''', (order_id,))
    assigned_mechanics = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(assigned_mechanics)


@app.route('/get_active_mechanics/<order_id>', methods=['GET'])
def get_active_mechanics(order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute('''
        SELECT m.Name, t.StartTime
        FROM Mechanics m
        JOIN TimeTracking t ON m.MechanicID = t.MechanicID
        WHERE t.OrderID = %s AND t.EndTime IS NULL
    ''', (order_id,))
    active_mechanics = cursor.fetchall()

    for mechanic in active_mechanics:
        mechanic['StartTime'] = (mechanic['StartTime'] + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')

    cursor.close()
    conn.close()

    return jsonify(active_mechanics)


def get_order_data(company):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    client_vehicle_table = 'VehiclesClientes'

    cursor.execute('''
        SELECT wo.OrderID, wo.VehicleID, 
               IF(wo.ClienteCodigoSap IS NULL, v.VehicleName, vc.VehicleName) AS VehicleName, 
               wo.WorkType, wo.Status, wo.Lugar, 
               IF(wo.ClienteCodigoSap IS NULL, wo.Dueno, c.Nombre) AS Dueno, 
               wo.Marca, wo.CreatedTime
        FROM WorkOrders wo
        LEFT JOIN {} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN {} vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        LEFT JOIN Clientes c ON wo.ClienteCodigoSap = c.CodigoSap
        WHERE (wo.Empresa = %s OR wo.ClienteCodigoSap IN (SELECT CodigoSap FROM Clientes))
        AND wo.Status != 'Completed'
        AND EXISTS (
            SELECT 1 FROM MechanicWorkOrder mwo WHERE mwo.OrderID = wo.OrderID
        )
        ORDER BY VehicleName
    '''.format(vehicle_table_map[company], client_vehicle_table), (company,))
    orders_with_mechanics = cursor.fetchall()

    for order in orders_with_mechanics:
        # Obtener mecánicos asignados
        cursor.execute('''
            SELECT DISTINCT m.Name
            FROM Mechanics m
            JOIN TimeTracking t ON m.MechanicID = t.MechanicID
            WHERE t.OrderID = %s
        ''', (order['OrderID'],))
        order['assigned_mechanics'] = [m['Name'] for m in cursor.fetchall()]

        # Obtener mecánicos activos
        cursor.execute('''
            SELECT m.Name, t.StartTime
            FROM Mechanics m
            JOIN TimeTracking t ON m.MechanicID = t.MechanicID
            WHERE t.OrderID = %s AND t.EndTime IS NULL
        ''', (order['OrderID'],))
        active_mechanics = cursor.fetchall()
        order['active_mechanics'] = [
            {'Name': m['Name'], 'StartTime': (m['StartTime'] + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')}
            for m in active_mechanics
        ]

        if len(order['active_mechanics']) > 0:
            order['status'] = 'green'
        elif 'Listo' in [s['Status'] for s in order.get('supplies_ready', [])]:
            order['status'] = 'red'
        else:
            order['status'] = 'yellow'

    cursor.execute('''
        SELECT wo.OrderID, wo.VehicleID, 
               IF(wo.ClienteCodigoSap IS NULL, v.VehicleName, vc.VehicleName) AS VehicleName, 
               wo.WorkType, wo.Status, wo.Lugar, 
               IF(wo.ClienteCodigoSap IS NULL, wo.Dueno, c.Nombre) AS Dueno, 
               wo.Marca, wo.CreatedTime, s.Description AS SupplyDescription, wos.Status AS SupplyStatus
        FROM WorkOrders wo
        LEFT JOIN {} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN {} vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        LEFT JOIN Clientes c ON wo.ClienteCodigoSap = c.CodigoSap
        LEFT JOIN WorkOrderSupplies wos ON wo.OrderID = wos.OrderID
        LEFT JOIN Supplies s ON wos.CodigoSap = s.CodigoSap
        WHERE (wo.Empresa = %s OR wo.ClienteCodigoSap IN (SELECT CodigoSap FROM Clientes))
        AND wos.Status = 'Listo'
    '''.format(vehicle_table_map[company], client_vehicle_table), (company,))
    orders_with_supplies_ready = cursor.fetchall()

    cursor.close()
    conn.close()

    return {
        'orders_with_mechanics': orders_with_mechanics,
        'orders_with_supplies_ready': orders_with_supplies_ready
    }

def refresh_orders_for_company(company):
    requests.get(f'http://127.0.0.1:5000/get_order_data/{company}')

@app.route('/get_order_data/<company>', methods=['GET'])
def get_order_data_api(company):
    data = get_order_data(company)
    return jsonify(data)



#clientes y sus vehiculos de sap
def refresh_clients_from_sap():
    hana_conn = connect_hana()
    if hana_conn:
        try:
            hana_cursor = hana_conn.cursor()
            # Execute the query to fetch clients with the additional filter
            hana_cursor.execute('''
                SELECT T0."CardCode", T0."CardName" 
                FROM "SBO_MONTASAHN".OCRD T0 
                WHERE T0."validFor" = 'Y' AND T0."CardType" = 'C'
            ''')
            clients_from_sap = hana_cursor.fetchall()

            # Insert clients data into MySQL
            local_conn = get_db_connection()
            local_cursor = local_conn.cursor()

            for codigo_sap, nombre in clients_from_sap:
                # Handle None values for nombre
                if nombre is None:
                    nombre = None

                local_cursor.execute('SELECT * FROM Clientes WHERE CodigoSap = %s', (codigo_sap,))
                existing_client = local_cursor.fetchone()

                if existing_client:
                    local_cursor.execute('''
                        UPDATE Clientes
                        SET Nombre = %s
                        WHERE CodigoSap = %s
                    ''', (nombre, codigo_sap))
                else:
                    local_cursor.execute('''
                        INSERT INTO Clientes (CodigoSap, Nombre)
                        VALUES (%s, %s)
                    ''', (codigo_sap, nombre))

            local_conn.commit()
            local_cursor.close()
            local_conn.close()

        except hana.Error as e:
            print(f"Error fetching data from SAP HANA: {e}")
        finally:
            hana_cursor.close()
            hana_conn.close()
    else:
        print("Failed to connect to SAP HANA")





@app.route('/select_service_type/<company>', methods=['GET', 'POST'])
def select_service_type(company):
    if request.method == 'POST':
        service_type = request.form.get('service_type')
        if service_type == 'internal':
            return redirect(url_for('add_work_order', company=company))
        elif service_type == 'client':
            return redirect(url_for('add_client_work_order', company=company))

    return render_template('select_service_type.html', company=company)


@app.route('/add_client_work_order/<company>', methods=['GET', 'POST'])
def add_client_work_order(company):
    refresh_clients_from_sap()

    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'POST':
        cliente_codigo_sap = request.form.get('cliente_codigo_sap')
        if not cliente_codigo_sap:
            flash('Please select a client.', 'error')
            return redirect(url_for('add_client_work_order', company=company))

        vehicle_id = request.form.get('vehicle_id')
        if not vehicle_id:
            flash('Please select a vehicle.', 'error')
            cursor.execute('SELECT VehicleID, VehicleName FROM VehiclesClientes WHERE CodigoSap = %s AND Status = "Disponible"', (cliente_codigo_sap,))
            vehicles = cursor.fetchall()
            cursor.execute('SELECT CodigoSap, Nombre FROM Clientes')
            clientes = cursor.fetchall()
            return render_template('add_client_work_order.html', clientes=clientes, vehicles=vehicles, company=company, selected_cliente=cliente_codigo_sap)

        # Verificar si el VehicleID es válido
        cursor.execute('SELECT VehicleID FROM VehiclesClientes WHERE VehicleID = %s', (vehicle_id,))
        valid_vehicle = cursor.fetchone()
        if not valid_vehicle:
            flash('Invalid VehicleID. Please select a valid vehicle.', 'error')
            return redirect(url_for('add_client_work_order', company=company, cliente_codigo_sap=cliente_codigo_sap))

        work_type = request.form['work_type']
        description = request.form['description']

        cursor.execute('SELECT Nombre FROM Clientes WHERE CodigoSap = %s', (cliente_codigo_sap,))
        cliente_name = cursor.fetchone()[0]

        cursor.execute('SELECT VehicleName FROM VehiclesClientes WHERE VehicleID = %s', (vehicle_id,))
        vehicle_name = cursor.fetchone()[0]

        cursor.execute('''
            INSERT INTO WorkOrders (VehicleID, WorkType, Description, Status, Empresa, ClienteCodigoSap, Dueno)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (vehicle_id, work_type, description, 'Pending', company, cliente_codigo_sap, cliente_name))
        conn.commit()

        new_order_id = cursor.lastrowid  # Obtener el ID de la nueva orden de trabajo

        # Crear chequeos de calidad para la nueva orden de trabajo
        create_mechanic_quality_check(conn, new_order_id)
        create_logistics_quality_check(conn, new_order_id)

        cursor.close()
        conn.close()
        flash(f'Orden de trabajo para {vehicle_name} añadida exitosamente.', 'success')
        return redirect(url_for('work_orders', company=company))

    cursor.execute('SELECT CodigoSap, Nombre FROM Clientes')
    clientes = cursor.fetchall()

    vehicles = []
    cliente_codigo_sap = request.args.get('cliente_codigo_sap')
    if cliente_codigo_sap:
        cursor.execute('''
            SELECT vc.VehicleID, vc.VehicleName
            FROM VehiclesClientes vc
            WHERE vc.CodigoSap = %s AND vc.Status = "Disponible"
        ''', (cliente_codigo_sap,))
        vehicles = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('add_client_work_order.html', clientes=clientes, vehicles=vehicles, company=company, selected_cliente=cliente_codigo_sap)






















@app.route('/assign_supply/<int:order_id>', methods=['POST'])
def assign_supply(order_id):
    supply_code = request.form.get('supply_code')  # Usar el código del suministro como identificador
    quantity = request.form.get('quantity')
    default_status = "Esperando"

    if not supply_code or not quantity:
        flash('No supply or quantity selected!', 'error')
        return redirect(url_for('work_order_detail', order_id=order_id))

    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Obtener los detalles del suministro y el SupplyID de la base de datos local
        cursor.execute('''
            SELECT SupplyID, Description
            FROM Supplies
            WHERE CodigoSap = %s
        ''', (supply_code,))
        supply_details = cursor.fetchone()

        if supply_details:
            supply_id, description = supply_details  # Desempaqueta el SupplyID y la descripción

            # Insertar el suministro en WorkOrderSupplies usando los detalles obtenidos de la base de datos local
            cursor.execute('''
                INSERT INTO WorkOrderSupplies (OrderID, SupplyID, CodigoSap, Description, Quantity, Status)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (order_id, supply_id, supply_code, description, quantity, default_status))
            conn.commit()

            # Obtener los detalles del trabajo y el vehículo para actualizar los insumos correspondientes
            cursor.execute('''
                SELECT wo.VehicleID, wo.WorkType, wo.Empresa, wo.ClienteCodigoSap
                FROM WorkOrders wo
                WHERE wo.OrderID = %s
            ''', (order_id,))
            order_details = cursor.fetchone()
            if order_details:
                vehicle_id, work_type, company, cliente_codigo_sap = order_details

                if cliente_codigo_sap is None:
                    # Mapear la empresa a la tabla de vehículos correspondiente
                    vehicle_table_map = {
                        'MontasaHN': 'Vehicles',
                        'MontasaCR': 'vehiculosCR',
                        'Monhaco': 'vehiculosMonhaco'
                    }
                    vehicle_table = vehicle_table_map.get(company, 'Vehicles')
                else:
                    vehicle_table = 'VehiclesClientes'

                # Determinar la columna a actualizar basado en el tipo de trabajo
                column_to_update = 'InsumosMantenimiento' if work_type == 'Mantenimiento' else 'InsumosReparo'

                # Obtener la lista actual de insumos
                cursor.execute(f'SELECT {column_to_update} FROM {vehicle_table} WHERE VehicleID = %s', (vehicle_id,))
                current_insumos = cursor.fetchone()

                if current_insumos:
                    current_insumos = current_insumos[0]
                    if current_insumos:
                        insumos_list = current_insumos.split(',')
                        if supply_code not in insumos_list:
                            insumos_list.append(supply_code)
                            new_insumos = ','.join(insumos_list)
                        else:
                            new_insumos = current_insumos
                    else:
                        new_insumos = supply_code

                    # Actualizar la columna de insumos
                    cursor.execute(f'UPDATE {vehicle_table} SET {column_to_update} = %s WHERE VehicleID = %s', (new_insumos, vehicle_id))
                    conn.commit()
                    flash('Supply assigned successfully!', 'success')
                else:
                    flash('Failed to retrieve current supplies!', 'error')
            else:
                flash('Failed to retrieve order details!', 'error')
        else:
            flash('Supply details not found!', 'error')

    except Exception as err:
        print("Error: ", err)
        flash('Failed to assign supply!', 'error')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_detail', order_id=order_id))











@app.route('/unassign_supply/<int:order_id>/<supply_code>', methods=['POST'])
def unassign_supply(order_id, supply_code):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # Asegúrate de que 'CodigoSap' es el nombre correcto de la columna en tu tabla 'WorkOrderSupplies'
        cursor.execute('DELETE FROM WorkOrderSupplies WHERE OrderID = %s AND CodigoSap = %s', (order_id, supply_code))
        conn.commit()
        flash('Supply unassigned successfully!', 'success')
    except Exception as err:
        print("Error: ", err)
        flash('Failed to unassign supply!', 'error')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_detail', order_id=order_id))


def refresh_supplies_from_sap(codigo_sap):
    # Conectar a SAP HANA
    hana_conn = connect_hana()
    if hana_conn:
        try:
            hana_cursor = hana_conn.cursor()
            # Ejecutar la consulta en SAP HANA
            query = f'''
                SELECT T0."ItemCode" as "CodigoSap", T0."ItemName" as "Description", T1."OnHand" as "QuantityInStock"
                FROM SBO_MONTASAHN.OITM T0
                LEFT JOIN SBO_MONTASAHN.OITW T1 ON T0."ItemCode" = T1."ItemCode"
                WHERE T1."WhsCode" = '04' AND T0."validFor" = 'Y' AND T0."ItemCode" = '{codigo_sap}'
            '''
            hana_cursor.execute(query)
            sap_supply = hana_cursor.fetchone()
            hana_cursor.close()
            hana_conn.close()

            if sap_supply:
                # Conectar a la base de datos local (SQL)
                local_conn = get_db_connection()
                local_cursor = local_conn.cursor()

                local_cursor.execute('SELECT * FROM Supplies WHERE CodigoSap = %s', (codigo_sap,))
                existing_supply = local_cursor.fetchone()

                if existing_supply:
                    local_cursor.close()
                    local_conn.close()
                    return False  # El suministro ya existe en la base de datos local
                else:
                    # Insertar el nuevo suministro si no existe
                    local_cursor.execute('''
                        INSERT INTO Supplies (CodigoSap, Description, QuantityInStock)
                        VALUES (%s, %s, %s)
                    ''', (sap_supply[0], sap_supply[1], sap_supply[2]))
                    local_conn.commit()
                    local_cursor.close()
                    local_conn.close()
                    return True
            else:
                return None  # El suministro no se encontró en SAP HANA

        except hana.Error as e:
            print(f"Error al obtener suministro de SAP HANA: {e}")
            return None
    else:
        return None  # Falló la conexión a SAP HANA




@app.route('/check_supply/<codigo_sap>', methods=['GET'])
def check_supply(codigo_sap):
    # Conectar a la base de datos local (SQL)
    local_conn = get_db_connection()
    local_cursor = local_conn.cursor(dictionary=True)

    # Verificar si el suministro ya existe en la base de datos local
    local_cursor.execute('SELECT Description FROM Supplies WHERE CodigoSap = %s', (codigo_sap,))
    existing_supply = local_cursor.fetchone()

    if existing_supply:
        local_cursor.close()
        local_conn.close()
        return jsonify({'exists': True, 'description': existing_supply['Description']})

    # Si no existe, verificar en SAP HANA
    hana_conn = connect_hana()
    if hana_conn:
        try:
            hana_cursor = hana_conn.cursor()
            query = f'''
                SELECT T0."ItemCode" as "CodigoSap", T0."ItemName" as "Description", T1."OnHand" as "QuantityInStock"
                FROM SBO_MONTASAHN.OITM T0
                LEFT JOIN SBO_MONTASAHN.OITW T1 ON T0."ItemCode" = T1."ItemCode"
                WHERE T1."WhsCode" = '04' AND T0."validFor" = 'Y' AND T0."ItemCode" = '{codigo_sap}'
            '''
            hana_cursor.execute(query)
            sap_supply = hana_cursor.fetchone()
            hana_cursor.close()
            hana_conn.close()

            if sap_supply:
                return jsonify({'exists': False, 'inSAP': True, 'description': sap_supply[1]})
            else:
                return jsonify({'exists': False, 'inSAP': False})

        except hana.Error as e:
            print(f"Error fetching supply from SAP HANA: {e}")
            return jsonify({'exists': False, 'inSAP': False, 'error': str(e)})
    else:
        return jsonify({'exists': False, 'inSAP': False, 'error': 'Failed to connect to SAP HANA'})


# Asegúrate de actualizar también la ruta original de 'refresh_supplies'
@app.route('/refresh_supplies/<company>', methods=['GET', 'POST'])
def refresh_supplies(company):
    if request.method == 'POST':
        codigo_sap = request.form.get('codigo_sap')
        if codigo_sap:
            result = refresh_supplies_from_sap(codigo_sap)
            if result is True:
                flash('Suministro añadido exitosamente.', 'success')
            elif result is False:
                flash('El suministro ya existe en la base de datos.', 'error')
            else:
                flash('Suministro no encontrado en SAP HANA.', 'error')
        else:
            flash('Por favor, ingrese un Código SAP.', 'error')
        return redirect(url_for('refresh_supplies', company=company))

    return render_template('refresh_supplies.html', company=company)








#supply status update
@app.route('/supply_status_update/<company>', methods=['GET', 'POST'])
def supply_status_update(company):
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        

        if 'new_status' in request.form:
            order_id = request.form['order_id']
            supply_id = request.form['supply_id']
            new_status = request.form['new_status']

            
            if new_status == 'Recibido':
                cursor.execute('''
                    UPDATE WorkOrderSupplies
                    SET Status = %s, ReceivedDate = NOW()
                    WHERE OrderID = %s AND SupplyID = %s
                ''', (new_status, order_id, supply_id))
            else:
                cursor.execute('''
                    UPDATE WorkOrderSupplies
                    SET Status = %s
                    WHERE OrderID = %s AND SupplyID = %s
                ''', (new_status, order_id, supply_id))
            conn.commit()

        if 'mechanic_pin' in request.form:
            mechanic_pin = request.form['mechanic_pin']
            checked_supplies = request.form.getlist('supply_ids[]')
            logging.debug(f'Mechanic PIN: {mechanic_pin}, Supplies: {checked_supplies}')
            
            cursor.execute('SELECT MechanicID FROM Mechanics WHERE PinCode = %s', (mechanic_pin,))
            mechanic = cursor.fetchone()

            if mechanic:
                mechanic_id = mechanic['MechanicID']
                for supply_id in checked_supplies:
                    logging.debug(f'Updating SupplyID: {supply_id} with MechanicID: {mechanic_id}')
                    cursor.execute('''
                        UPDATE WorkOrderSupplies
                        SET Status = 'Recibido', ReceivedByMechanicID = %s, ReceivedDate = NOW()
                        WHERE SupplyID = %s
                    ''', (mechanic_id, supply_id))
                conn.commit()
                flash('Suministros marcados como recibidos.', 'success')
            else:
                flash('PIN de mecánico inválido.', 'error')

        if 'boss_pin' in request.form:
            boss_pin = request.form['boss_pin']
            checked_supplies = request.form.getlist('boss_supply_ids[]')
            logging.debug(f'Boss PIN: {boss_pin}, Supplies: {checked_supplies}')
            
            cursor.execute('SELECT BossID FROM Boss WHERE BossPin = %s', (boss_pin,))
            boss = cursor.fetchone()

            if boss:
                boss_id = boss['BossID']
                for supply_id in checked_supplies:
                    logging.debug(f'Approving SupplyID: {supply_id} with BossID: {boss_id}')
                    cursor.execute('''
                        UPDATE WorkOrderSupplies
                        SET ApprovedByBossID = %s
                        WHERE SupplyID = %s
                    ''', (boss_id, supply_id))
                conn.commit()
                flash('Suministros aprobados.', 'success')
            else:
                flash('PIN de jefe inválido.', 'error')

    search_order_id = request.args.get('search_order_id', '')
    search_codigo_sap = request.args.get('search_codigo_sap', '')
    search_name = request.args.get('search_name', '')
    offset = int(request.args.get('offset', 0))
    limit = 20

    query = '''
        SELECT ws.OrderID, ws.SupplyID, s.CodigoSap, s.Description, ws.Quantity, ws.Status, ws.ReceivedByMechanicID, m.Name as MechanicName, b.Nombre as BossName, ws.ReceivedDate
        FROM WorkOrderSupplies ws
        JOIN Supplies s ON ws.SupplyID = s.SupplyID
        LEFT JOIN Mechanics m ON ws.ReceivedByMechanicID = m.MechanicID
        LEFT JOIN Boss b ON ws.ApprovedByBossID = b.BossID
        WHERE (%s = '' OR ws.OrderID LIKE %s)
        AND (%s = '' OR s.CodigoSap LIKE %s)
        AND (%s = '' OR s.Description LIKE %s)
        ORDER BY CASE WHEN ws.Status = 'Listo' THEN 1 WHEN ws.Status = 'Esperando' THEN 2 ELSE 3 END, ws.OrderID DESC
        LIMIT %s OFFSET %s
    '''

    cursor.execute(query, (
        search_order_id, f'%{search_order_id}%',
        search_codigo_sap, f'%{search_codigo_sap}%',
        search_name, f'%{search_name}%',
        limit, offset
    ))

    supplies = cursor.fetchall()
    
    cursor.execute('''
        SELECT COUNT(*) as total
        FROM WorkOrderSupplies ws
        JOIN Supplies s ON ws.SupplyID = s.SupplyID
        WHERE (%s = '' OR ws.OrderID LIKE %s)
        AND (%s = '' OR s.CodigoSap LIKE %s)
        AND (%s = '' OR s.Description LIKE %s)
    ''', (
        search_order_id, f'%{search_order_id}%',
        search_codigo_sap, f'%{search_codigo_sap}%',
        search_name, f'%{search_name}%'
    ))

    total_supplies = cursor.fetchone()['total']
    total_pages = (total_supplies + limit - 1) // limit

    cursor.close()
    conn.close()
    
    return render_template('supply_status_update.html', supplies=supplies, company=company, offset=offset, limit=limit, total_pages=total_pages, current_page=(offset // limit) + 1)


#supply status update

#mechanic work hours summary

def get_week_range(date):
    # Convert to a datetime object if the input is a string
    if isinstance(date, str):
        date = datetime.strptime(date, "%Y-%m-%d").date()

    # Find the most recent Monday
    week_start = date - timedelta(days=date.weekday())

    # Find the following Sunday
    week_end = week_start + timedelta(days=6)

    return week_start, week_end

def update_mechanic_work_hour_summary(mechanic_id, date):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Calculate work hours for the day as TIME
        cursor.execute('''
            SELECT SEC_TO_TIME(SUM(TIME_TO_SEC(TIMEDIFF(EndTime, StartTime))))
            FROM TimeTracking
            WHERE MechanicID = %s AND DATE(StartTime) = %s
        ''', (mechanic_id, date))
        daily_duration = cursor.fetchone()[0] or '00:00:00'

        # Update or insert into DailyMechanicWorkHours
        cursor.execute('''
            INSERT INTO DailyMechanicWorkHours (MechanicID, Date, TotalHours)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE TotalHours = VALUES(TotalHours)
        ''', (mechanic_id, date, daily_duration))

        # Calculate work hours for the week
        week_start, week_end = get_week_range(date)
        cursor.execute('''
            SELECT SEC_TO_TIME(SUM(TIME_TO_SEC(TIMEDIFF(EndTime, StartTime))))
            FROM TimeTracking
            WHERE MechanicID = %s AND DATE(StartTime) BETWEEN %s AND %s
        ''', (mechanic_id, week_start, week_end))
        weekly_duration = cursor.fetchone()[0] or '00:00:00'

        # Update or insert into WeeklyMechanicWorkHours
        cursor.execute('''
            INSERT INTO WeeklyMechanicWorkHours (MechanicID, WeekStartDate, TotalHours)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE TotalHours = VALUES(TotalHours)
        ''', (mechanic_id, week_start, weekly_duration))

        # Similarly, calculate work hours for the month and update MonthlyMechanicWorkHours
        month_start = date.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cursor.execute('''
            SELECT SEC_TO_TIME(SUM(TIME_TO_SEC(TIMEDIFF(EndTime, StartTime))))
            FROM TimeTracking
            WHERE MechanicID = %s AND DATE(StartTime) BETWEEN %s AND %s
        ''', (mechanic_id, month_start, month_end))
        monthly_duration = cursor.fetchone()[0] or '00:00:00'

        month_str = date.strftime('%Y-%m')  # Format as 'YYYY-MM'
        cursor.execute('''
            INSERT INTO MonthlyMechanicWorkHours (MechanicID, Month, TotalHours)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE TotalHours = VALUES(TotalHours)
        ''', (mechanic_id, month_str, monthly_duration))

        conn.commit()
    except mysql.connector.Error as err:
        print("SQL Error: ", err)
    finally:
        cursor.close()
        conn.close()

def convert_to_hours(total_hours):
    if isinstance(total_hours, timedelta):
        total_seconds = int(total_hours.total_seconds())
        days = total_seconds // (24 * 3600)
        hours = (total_seconds % (24 * 3600)) // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        total_hours = days * 24 + hours
        return f"{total_hours:02}:{minutes:02}:{seconds:02}"
    else:
        # Regular expression to match days and time in string format
        match = re.match(r"((?P<days>\d+) day[s]?, )?(?P<hours>\d{1,2}):(?P<minutes>\d{2}):(?P<seconds>\d{2})", total_hours)
        if not match:
            return total_hours  # Return as is if the format is unexpected

        time_data = match.groupdict()
        days = int(time_data.get('days') or 0)
        hours = int(time_data['hours'])
        minutes = int(time_data['minutes'])
        seconds = int(time_data['seconds'])

        total_hours = days * 24 + hours
        return f"{total_hours:02}:{minutes:02}:{seconds:02}"

@app.route('/mechanic_work_hours_summary/<company>', methods=['GET', 'POST'])
def mechanic_work_hours_summary(company):
    filter_date_str = request.args.get('filter_date', datetime.now().strftime('%Y-%m-%d'))
    filter_date = datetime.strptime(filter_date_str, '%Y-%m-%d')

    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        # Fetch data for daily summary
        cursor.execute('''
            SELECT d.MechanicID, m.Name as MechanicName, d.Date, d.TotalHours 
            FROM DailyMechanicWorkHours d
            JOIN Mechanics m ON d.MechanicID = m.MechanicID
            WHERE d.Date = %s
        ''', (filter_date_str,))
        daily_summary = cursor.fetchall()
        for entry in daily_summary:
            entry['TotalHours'] = convert_to_hours(entry['TotalHours'])

        # Calculate week range
        week_start, week_end = get_week_range(filter_date_str)
        cursor.execute('''
            SELECT w.MechanicID, m.Name as MechanicName, w.WeekStartDate, w.TotalHours 
            FROM WeeklyMechanicWorkHours w
            JOIN Mechanics m ON w.MechanicID = m.MechanicID
            WHERE w.WeekStartDate = %s
        ''', (week_start,))
        weekly_summary = cursor.fetchall()
        for entry in weekly_summary:
            entry['TotalHours'] = convert_to_hours(entry['TotalHours'])

        # Fetch data for monthly summary
        month_str = filter_date.strftime("%Y-%m")  # Format month and year as "YYYY-MM"
        cursor.execute('''
            SELECT mo.MechanicID, m.Name as MechanicName, mo.Month, mo.TotalHours 
            FROM MonthlyMechanicWorkHours mo
            JOIN Mechanics m ON mo.MechanicID = m.MechanicID
            WHERE mo.Month = %s
        ''', (month_str,))
        monthly_summary = cursor.fetchall()
        for entry in monthly_summary:
            entry['TotalHours'] = convert_to_hours(entry['TotalHours'])

    finally:
        cursor.close()
        conn.close()

    return render_template('mechanic_work_hours_summary.html', 
                           daily_summary=daily_summary, 
                           weekly_summary=weekly_summary, 
                           monthly_summary=monthly_summary, 
                           filter_date=filter_date_str, 
                           company=company)
#mechanic work hours summary


#unassign mechanic

@app.route('/assign_mechanic', methods=['POST'])
def assign_mechanic():
    order_id = request.form['order_id']
    mechanic_id = request.form['mechanic_id']
    start_time = datetime.now()

    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Verificar si el mecánico ya está asignado a cualquier orden
        cursor.execute('SELECT COUNT(*) FROM MechanicWorkOrder WHERE MechanicID = %s', (mechanic_id,))
        if cursor.fetchone()[0] > 0:
            flash('El mecánico ya está asignado a una orden de trabajo.', 'error')
            return redirect(url_for('work_order_detail', order_id=order_id))

        # Verificar si la orden existe
        cursor.execute('SELECT VehicleID FROM WorkOrders WHERE OrderID = %s', (order_id,))
        vehicle_id = cursor.fetchone()
        if not vehicle_id:
            return 'Order not found', 404
        vehicle_id = vehicle_id[0]

        # Asignar mecánico a la orden
        cursor.execute('REPLACE INTO MechanicWorkOrder (OrderID, MechanicID) VALUES (%s, %s)', (order_id, mechanic_id))

        # Insertar en TimeTracking
        cursor.execute('INSERT INTO TimeTracking (OrderID, MechanicID, StartTime, VehicleID) VALUES (%s, %s, %s, %s)', (order_id, mechanic_id, start_time, vehicle_id))

        conn.commit()

        update_mechanic_work_hour_summary(mechanic_id, start_time.date())
    except mysql.connector.Error as err:
        conn.rollback()
        flash(str(err), 'error')
        return redirect(url_for('work_order_detail', order_id=order_id))
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_detail', order_id=order_id))



def unassign_mechanic_logic(order_id, mechanic_id):
    end_time = datetime.now()
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        cursor.execute('UPDATE TimeTracking SET EndTime = %s WHERE OrderID = %s AND MechanicID = %s AND EndTime IS NULL', (end_time, order_id, mechanic_id))

        cursor.execute('DELETE FROM MechanicWorkOrder WHERE OrderID = %s AND MechanicID = %s', (order_id, mechanic_id))

        conn.commit()

        update_mechanic_work_hour_summary(mechanic_id, end_time.date())
    except mysql.connector.Error as err:
        conn.rollback()
        print(f"Error unassigning mechanic: {err}")
    finally:
        cursor.close()
        conn.close()




@app.route('/unassign_mechanic', methods=['POST'])
def unassign_mechanic():
    order_id = request.form['order_id']
    mechanic_id = request.form['mechanic_id']

    try:
        unassign_mechanic_logic(order_id, mechanic_id)
        flash('Mechanic unassigned successfully!', 'success')
    except Exception as e:
        flash(f'Error unassigning mechanic: {str(e)}', 'error')

    return redirect(url_for('work_order_detail', order_id=order_id))



#assign mechanic




#
#vehiculos

@app.route('/vehicles')
def vehicles_list():
    company = request.args.get('company', 'MontasaHN')  # Default to MontasaHN if no company is specified
    vehicle_id = request.args.get('vehicle_id')
    availability = request.args.get('availability')
    conn = get_db_connection()
    cursor = conn.cursor()

    # Seleccionar la tabla según la empresa
    if company == 'MontasaHN':
        table_name = 'Vehicles'
    elif company == 'MontasaCR':
        table_name = 'VehiculosCR'
    elif company == 'Monhaco':
        table_name = 'VehiculosMonhaco'
    else:
        table_name = 'Vehicles'  # Default table if no valid company is specified

    # Fetch all vehicles for the dropdown
    dropdown_query = f"SELECT * FROM {table_name} WHERE Status IN ('Disponible', 'En Renta', 'En Taller', 'Reparación Externa')"
    cursor.execute(dropdown_query)
    all_vehicles = cursor.fetchall()

    # Fetch filtered vehicles based on selected criteria
    query = f"SELECT * FROM {table_name}"
    conditions = []
    parameters = []

    if vehicle_id:
        conditions.append("VehicleID = %s")
        parameters.append(vehicle_id)
    if availability:
        conditions.append("Status = %s")
        parameters.append(availability)

    # Excluir "Deshabilitado" y "Vendido" por defecto
    if not availability:
        conditions.append("Status NOT IN ('Deshabilitado', 'Vendido')")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    # Añadir la cláusula ORDER BY para ordenar los vehículos por HorometroDesdeUltimoMantenimiento de mayor a menor
    query += " ORDER BY HorometroDesdeUltimoMantenimiento DESC"

    cursor.execute(query, tuple(parameters))
    vehicles = cursor.fetchall()

    # Consultas para contar los vehículos por estado, ajustadas a la tabla específica de la empresa
    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE Status = 'Disponible'")
    disponible_count = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE Status = 'En Renta'")
    en_renta_count = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE Status = 'En Taller'")
    en_taller_count = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE Status = 'Reparación Externa'")
    reparacion_externa_count = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE Status = 'Deshabilitado'")
    deshabilitado_count = cursor.fetchone()[0]

    cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE Status = 'Vendido'")
    vendido_count = cursor.fetchone()[0]

    total_vehicles = disponible_count + en_renta_count + en_taller_count + reparacion_externa_count + deshabilitado_count + vendido_count
    disponible_percentage = (disponible_count / total_vehicles) * 100 if total_vehicles > 0 else 0
    en_renta_percentage = (en_renta_count / total_vehicles) * 100 if total_vehicles > 0 else 0
    en_taller_percentage = (en_taller_count / total_vehicles) * 100 if total_vehicles > 0 else 0
    reparacion_externa_percentage = (reparacion_externa_count / total_vehicles) * 100 if total_vehicles > 0 else 0
    deshabilitado_percentage = (deshabilitado_count / total_vehicles) * 100 if total_vehicles > 0 else 0
    vendido_percentage = (vendido_count / total_vehicles) * 100 if total_vehicles > 0 else 0

    cursor.close()
    conn.close()

    # Pasar también la empresa seleccionada a la plantilla para mostrarla en la interfaz
    return render_template('vehicles.html', company=company, vehicles=vehicles, disponible_count=disponible_count, en_renta_count=en_renta_count, en_taller_count=en_taller_count, disponible_percentage=disponible_percentage, en_renta_percentage=en_renta_percentage, en_taller_percentage=en_taller_percentage, reparacion_externa_count=reparacion_externa_count, reparacion_externa_percentage=reparacion_externa_percentage, deshabilitado_count=deshabilitado_count, deshabilitado_percentage=deshabilitado_percentage, vendido_count=vendido_count, vendido_percentage=vendido_percentage, vehicle_id=vehicle_id, all_vehicles=all_vehicles)



def get_vehicle_details(company, vehicle_name):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    query = f"""
    SELECT 
        Type, Marca, Modelo, Ano, Serie, VehicleName, Observacion, CapacidadKG, Status,
        Horometro, HorometroDesdeUltimoMantenimiento, FechaActualizacionHorometro, Ubicacion, Empresa,
        `Capacidad Maxima`, `Altura Max. Elevado`, `Peso Bruto`, Ancho, Largo, Altura, Motor, `Tipo de Llantas`,
        `Horquillas Estándar`, Extintor, `Luz Estroboscopica`, `Luces de Trabajo delantera, trasera, vias`, Retrovisores,
        `Asiento Ergonomico`, `Cinturon de Seguridad`, Capote
    FROM {vehicle_table}
    WHERE VehicleName = %s AND Empresa = %s
    """
    cursor.execute(query, (vehicle_name, company))
    vehicle_details = cursor.fetchone()
    
    cursor.close()
    conn.close()
    
    if vehicle_details:
        return vehicle_details
    else:
        return {
            'Type': '',
            'Marca': '',
            'Modelo': '',
            'Ano': '',
            'Serie': '',
            'VehicleName': '',
            'Observacion': '',
            'CapacidadKG': '',
            'Status': '',
            'Horometro': '',
            'HorometroDesdeUltimoMantenimiento': '',
            'FechaActualizacionHorometro': '',
            'Ubicacion': '',
            'Empresa': '',
            'Capacidad Maxima': '',
            'Altura Max. Elevado': '',
            'Peso Bruto': '',
            'Ancho': '',
            'Largo': '',
            'Altura': '',
            'Motor': '',
            'Tipo de Llantas': '',
            'Horquillas Estándar': '',
            'Extintor': '',
            'Luz Estroboscopica': '',
            'Luces de Trabajo delantera, trasera, vias': '',
            'Retrovisores': '',
            'Asiento Ergonomico': '',
            'Cinturon de Seguridad': '',
            'Capote': ''
        }


@app.route('/vehicle_detail/<company>/<vehicle_name>', methods=['GET', 'POST'])
def vehicle_detail(company, vehicle_name):
    error_message = request.args.get('error_message')
    vehicle_details = get_vehicle_details(company, vehicle_name)

    if request.method == 'POST':
        vehicle_details = {
            'Type': request.form['type'],
            'Marca': request.form['marca'],
            'Modelo': request.form['modelo'],
            'Ano': request.form['ano'],
            'Serie': request.form['serie'],
            'VehicleName': request.form['vehicle_name'],
            'Observacion': request.form['observacion'],
            'CapacidadKG': request.form['capacidadKG'],
            'Status': request.form['status'],
            'Horometro': request.form['horometro'],
            'HorometroDesdeUltimoMantenimiento': request.form['horometro_desde_ultimo_mantenimiento'],
            'FechaActualizacionHorometro': request.form['fecha_actualizacion_horometro'],
            'Ubicacion': request.form['ubicacion'],
            'Empresa': request.form['empresa'],
            'Capacidad Maxima': request.form['capacidad_maxima'],
            'Altura Max. Elevado': request.form['altura_max_elevado'],
            'Peso Bruto': request.form['peso_bruto'],
            'Ancho': request.form['ancho'],
            'Largo': request.form['largo'],
            'Altura': request.form['altura'],
            'Motor': request.form['motor'],
            'Tipo de Llantas': request.form['tipo_llantas'],
            'Horquillas Estándar': request.form['horquillas_estandar'],
            'Extintor': 'extintor' in request.form,
            'Luz Estroboscopica': 'luz_estroboscopica' in request.form,
            'Luces de Trabajo delantera, trasera, vias': 'luces_trabajo' in request.form,
            'Retrovisores': 'retrovisores' in request.form,
            'Asiento Ergonomico': 'asiento_ergonomico' in request.form,
            'Cinturon de Seguridad': 'cinturon_seguridad' in request.form,
            'Capote': 'capote' in request.form
        }

        save_vehicle_details(company, vehicle_name, vehicle_details)

    return render_template('vehicle_detail.html', company=company, vehicle_name=vehicle_name, vehicle_details=vehicle_details, error_message=error_message)


def save_vehicle_details(company, vehicle_name, details):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    update_query = f"""
    UPDATE {vehicle_table}
    SET 
        Type = %s,
        Marca = %s,
        Modelo = %s,
        Ano = %s,
        Serie = %s,
        VehicleName = %s,
        Observacion = %s,
        CapacidadKG = %s,
        Status = %s,
        Horometro = %s,
        HorometroDesdeUltimoMantenimiento = %s,
        FechaActualizacionHorometro = %s,
        Ubicacion = %s,
        Empresa = %s,
        `Capacidad Maxima` = %s,
        `Altura Max. Elevado` = %s,
        `Peso Bruto` = %s,
        Ancho = %s,
        Largo = %s,
        Altura = %s,
        Motor = %s,
        `Tipo de Llantas` = %s,
        `Horquillas Estándar` = %s,
        Extintor = %s,
        `Luz Estroboscopica` = %s,
        `Luces de Trabajo delantera, trasera, vias` = %s,
        Retrovisores = %s,
        `Asiento Ergonomico` = %s,
        `Cinturon de Seguridad` = %s,
        Capote = %s
    WHERE 
        VehicleName = %s AND Empresa = %s;
    """
    
    cursor.execute(update_query, (
        details['Type'],
        details['Marca'],
        details['Modelo'],
        details['Ano'],
        details['Serie'],
        details['VehicleName'],
        details['Observacion'],
        details['CapacidadKG'],
        details['Status'],
        details['Horometro'],
        details['HorometroDesdeUltimoMantenimiento'],
        details['FechaActualizacionHorometro'],
        details['Ubicacion'],
        details['Empresa'],
        details['Capacidad Maxima'],
        details['Altura Max. Elevado'],
        details['Peso Bruto'],
        details['Ancho'],
        details['Largo'],
        details['Altura'],
        details['Motor'],
        details['Tipo de Llantas'],
        details['Horquillas Estándar'],
        details['Extintor'],
        details['Luz Estroboscopica'],
        details['Luces de Trabajo delantera, trasera, vias'],
        details['Retrovisores'],
        details['Asiento Ergonomico'],
        details['Cinturon de Seguridad'],
        details['Capote'],
        vehicle_name,
        company
    ))
    
    conn.commit()
    cursor.close()
    conn.close()

@app.route('/save_vehicle_details/<company>/<vehicle_name>', methods=['POST'])
def save_vehicle_details_route(company, vehicle_name):
    vehicle_details = {
        'Type': request.form.get('type', ''),
        'Marca': request.form.get('marca', ''),
        'Modelo': request.form.get('modelo', ''),
        'Ano': request.form.get('ano', ''),
        'Serie': request.form.get('serie', ''),
        'VehicleName': request.form.get('vehicle_name', ''),
        'Observacion': request.form.get('observacion', ''),
        'CapacidadKG': request.form.get('capacidadKG', ''),
        'Status': request.form.get('status', ''),
        'Horometro': request.form.get('horometro', ''),
        'HorometroDesdeUltimoMantenimiento': request.form.get('horometro_desde_ultimo_mantenimiento', ''),
        'FechaActualizacionHorometro': request.form.get('fecha_actualizacion_horometro', ''),
        'Ubicacion': request.form.get('ubicacion', ''),
        'Empresa': request.form.get('empresa', ''),
        'Capacidad Maxima': request.form.get('capacidad_maxima', ''),
        'Altura Max. Elevado': request.form.get('altura_max_elevado', ''),
        'Peso Bruto': request.form.get('peso_bruto', ''),
        'Ancho': request.form.get('ancho', ''),
        'Largo': request.form.get('largo', ''),
        'Altura': request.form.get('altura', ''),
        'Motor': request.form.get('motor', ''),
        'Tipo de Llantas': request.form.get('tipo_llantas', ''),
        'Horquillas Estándar': request.form.get('horquillas_estandar', ''),
        'Extintor': 'extintor' in request.form,
        'Luz Estroboscopica': 'luz_estroboscopica' in request.form,
        'Luces de Trabajo delantera, trasera, vias': 'luces_trabajo' in request.form,
        'Retrovisores': 'retrovisores' in request.form,
        'Asiento Ergonomico': 'asiento_ergonomico' in request.form,
        'Cinturon de Seguridad': 'cinturon_seguridad' in request.form,
        'Capote': 'capote' in request.form
    }

    save_vehicle_details(company, vehicle_name, vehicle_details)

    return redirect(url_for('vehicle_detail', company=company, vehicle_name=vehicle_name))







@app.route('/vehicles_client_list')
def vehicles_client_list():
    company = request.args.get('company', 'MontasaHN')
    client_id = request.args.get('client_id', None)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT CodigoSap, Nombre FROM Clientes")
    clients = cursor.fetchall()

    vehicles = []
    if client_id:
        cursor.execute("SELECT * FROM VehiclesClientes WHERE CodigoSap = %s", (client_id,))
        vehicles = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('vehicles_client_list.html', company=company, clients=clients, vehicles=vehicles)


@app.route('/get_vehicles/<client_id>')
def get_vehicles(client_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM VehiclesClientes WHERE CodigoSap = %s", (client_id,))
    vehicles = cursor.fetchall()
    app.logger.debug(f"Vehículos filtrados: {vehicles}")

    cursor.close()
    conn.close()

    return jsonify(vehicles)

@app.route('/filter_clients')
def filter_clients():
    search_query = request.args.get('query', '')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT CodigoSap, Nombre FROM Clientes WHERE CodigoSap LIKE %s OR Nombre LIKE %s", (f'%{search_query}%', f'%{search_query}%'))
    clients = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(clients)

#
#vehiculos















#Ordenes de SERVICIO Principio

@app.route('/add_work_order/<company>', methods=['GET', 'POST'])
def add_work_order(company):
    conn = get_db_connection()
    cursor = conn.cursor()

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }

    # Obtén la tabla de vehículos correcta según la compañía, usa 'Vehicles' como predeterminado
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Construye y ejecuta la consulta SQL utilizando la tabla de vehículos correcta
    sql_query = '''
        SELECT VehicleID, VehicleName, Marca, Empresa FROM {}
        WHERE Status IN ("Disponible", "En Renta") OR VehicleName = "Generico"
    '''.format(vehicle_table)  # Usa format para insertar el nombre de la tabla

    cursor.execute(sql_query)
    vehicles = cursor.fetchall()

    if request.method == 'POST':
        # Extrae los valores del formulario
        vehicle_id = request.form['vehicle_id']
        work_type = request.form.get('work_type')
        lugar = request.form['lugar']
        description = request.form['description']

        # Recupera la marca y dueño del vehículo desde la tabla de vehículos
        cursor.execute('SELECT Type, Empresa FROM {} WHERE VehicleID = %s'.format(vehicle_table), (vehicle_id,))
        vehicle_data = cursor.fetchone()
        marca = vehicle_data[0]
        dueno = vehicle_data[1]

        # Check if the vehicle is generic
        cursor.execute('SELECT VehicleName FROM {} WHERE VehicleID = %s'.format(vehicle_table), (vehicle_id,))
        vehicle_name = cursor.fetchone()[0]

        # Allow multiple work orders for a generic vehicle
        if vehicle_name != "Generico":
            cursor.execute('SELECT * FROM WorkOrders WHERE VehicleID = %s AND Status != "Completed"', (vehicle_id,))
            existing_order = cursor.fetchone()

            if existing_order:
                message = "There is already an open work order for this vehicle."
                cursor.close()
                flash(message)
                return redirect(url_for('work_orders', company=company))

        # Determine the status based on the work_type
        work_order_status = "Maintenance" if work_type == 'Maintenance' else "Repair"

        # Insert the new work order
        cursor.execute('''
        INSERT INTO WorkOrders (VehicleID, WorkType, Description, Status, Lugar, Dueno, Marca, Empresa)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (vehicle_id, work_type, description, work_order_status, lugar, dueno, marca, company))
        conn.commit()
        new_order_id = cursor.lastrowid  # Get the ID of the newly created work order

        # Create mechanic and logistics quality checks for the new work order
        create_mechanic_quality_check(conn, new_order_id)
        create_logistics_quality_check(conn, new_order_id)

        # Update the vehicle's status and location based on the current status and work type
        cursor.execute('SELECT Status FROM {} WHERE VehicleID = %s'.format(vehicle_table), (vehicle_id,))
        current_status = cursor.fetchone()[0]
        vehicle_status = "En Taller" if current_status == "Disponible" else "Reparación Externa"

        cursor.execute('''
        UPDATE {} SET Status = %s, Ubicacion = %s WHERE VehicleID = %s
        '''.format(vehicle_table), (vehicle_status, lugar, vehicle_id))
        conn.commit()

        # Reset HorometroDesdeUltimoMantenimiento if the work type is Maintenance
        if work_type == 'Maintenance':
            cursor.execute('''
            UPDATE {} SET HorometroDesdeUltimoMantenimiento = 0 WHERE VehicleID = %s
            '''.format(vehicle_table), (vehicle_id,))
            conn.commit()

        cursor.close()
        flash('Work order added successfully.')
        return redirect(url_for('work_orders', company=company))

    cursor.close()
    return render_template('add_work_order.html', vehicles=vehicles, company=company)




# Add work order

@app.route('/work_orders/<company>')
def work_orders(company):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Filtra las órdenes de trabajo activas por empresa o cliente
    cursor.execute(f'''
        SELECT DISTINCT wo.OrderID
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID
        WHERE wo.Status != 'Completed' AND (wo.Empresa = %s OR wo.ClienteCodigoSap IN (SELECT CodigoSap FROM Clientes))
        ORDER BY wo.OrderID
    ''', (company,))
    active_order_ids = cursor.fetchall()

    # Filtra los nombres de vehículos activos por empresa o cliente
    cursor.execute(f'''
        SELECT DISTINCT IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE wo.Status != 'Completed' AND (wo.Empresa = %s OR wo.ClienteCodigoSap IN (SELECT CodigoSap FROM Clientes))
        ORDER BY VehicleName
    ''', (company,))
    active_vehicle_names = cursor.fetchall()

    # Retrieve all search terms and date range
    order_id_query = request.args.get('order_id', '')
    vehicle_name_query = request.args.get('vehicle_name', '')
    description_query = request.args.get('description', '')
    start_date = request.args.get('start_date', '2024-01-01')
    now = datetime.now()
    end_date = (now + timedelta(days=1)).strftime('%Y-%m-%d')
    end_date = request.args.get('end_date', end_date)

    # Get current page and set the number of items per page
    page = int(request.args.get('page', 1))
    items_per_page = 15
    offset = (page - 1) * items_per_page

    # Constructing the SQL query based on search parameters
    query_conditions = ["wo.Status != 'Completed'"]
    query_params = []

    if order_id_query:
        query_conditions.append("wo.OrderID LIKE %s")
        query_params.append(f'%{order_id_query}%')

    if vehicle_name_query:
        query_conditions.append(f"(v.VehicleName LIKE %s OR vc.VehicleName LIKE %s)")
        query_params.extend([f'%{vehicle_name_query}%', f'%{vehicle_name_query}%'])

    if description_query:
        query_conditions.append("wo.Description LIKE %s")
        query_params.append(f'%{description_query}%')

    query_conditions.append("wo.CreatedTime BETWEEN %s AND %s")
    query_params.extend([start_date, end_date])

    query_conditions.append("(wo.Empresa = %s OR wo.ClienteCodigoSap IN (SELECT CodigoSap FROM Clientes))")
    query_params.append(company)

    where_clause = " AND ".join(query_conditions)
    sql_query = f'''
        SELECT wo.OrderID, wo.VehicleID, wo.WorkType, wo.Description, IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName, wo.CreatedTime, wo.currently_waiting,
               (SELECT COUNT(*) FROM WorkOrderSupplies ws WHERE ws.OrderID = wo.OrderID AND ws.Status = 'Listo') AS ready_supplies_count,
               (SELECT COUNT(*) FROM MechanicWorkOrder mwo WHERE mwo.OrderID = wo.OrderID) AS active_mechanics_count
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE {where_clause}
        ORDER BY ready_supplies_count DESC, active_mechanics_count DESC, wo.OrderID DESC
        LIMIT %s OFFSET %s
    '''
    query_params.extend([items_per_page, offset])
    cursor.execute(sql_query, tuple(query_params))
    work_orders = cursor.fetchall()

    for order in work_orders:
        order_id = order['OrderID']

        # Fetch active mechanics for the order
        cursor.execute('''
            SELECT m.MechanicID, m.Name
            FROM Mechanics m
            JOIN MechanicWorkOrder mwo ON m.MechanicID = mwo.MechanicID
            WHERE mwo.OrderID = %s
        ''', (order_id,))
        order['active_mechanics'] = cursor.fetchall()

        # Fetch waiting supplies for the order
        cursor.execute('''
            SELECT ws.CodigoSap, s.Description, ws.Quantity
            FROM WorkOrderSupplies ws
            JOIN Supplies s ON ws.CodigoSap = s.CodigoSap
            WHERE ws.OrderID = %s AND ws.Status = 'Esperando'
        ''', (order_id,))
        order['waiting_supplies'] = cursor.fetchall()

        # Fetch ready supplies for the order
        cursor.execute('''
            SELECT ws.CodigoSap, s.Description, ws.Quantity
            FROM WorkOrderSupplies ws
            JOIN Supplies s ON ws.CodigoSap = s.CodigoSap
            WHERE ws.OrderID = %s AND ws.Status = 'Listo'
        ''', (order_id,))
        order['ready_supplies'] = cursor.fetchall()

    # Get the total number of work orders
    cursor.execute(f'''
        SELECT COUNT(*) as total
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE {where_clause}
    ''', tuple(query_params[:-2]))  # Exclude the LIMIT and OFFSET params
    total_orders = cursor.fetchone()['total']

    cursor.close()
    conn.close()

    total_pages = (total_orders + items_per_page - 1) // items_per_page

    return render_template('work_orders.html', 
                           work_orders=work_orders, 
                           active_order_ids=active_order_ids, 
                           active_vehicle_names=active_vehicle_names, 
                           order_id_query=order_id_query, 
                           vehicle_name_query=vehicle_name_query, 
                           description_query=description_query, 
                           start_date=start_date, 
                           end_date=end_date, 
                           company=company,
                           page=page,
                           total_pages=total_pages)




#work orders
#ruta con previously assigned supplies
@app.route('/work_order_detail/<int:order_id>', methods=['GET'])
def work_order_detail(order_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch the main details of the work order, handling both company and client vehicles
    cursor.execute('''
        SELECT wo.OrderID, wo.VehicleID, IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName, wo.WorkType, wo.Description, wo.Status,
            wo.Lugar, IFNULL(c.Nombre, wo.Dueno) AS Dueno, wo.Marca, wo.Diagnostico, wo.TrabajoRealizado, wo.CreatedTime, wo.Empresa, wo.currently_waiting
        FROM WorkOrders wo
        LEFT JOIN Vehicles v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        LEFT JOIN Clientes c ON wo.ClienteCodigoSap = c.CodigoSap
        WHERE wo.OrderID = %s
    ''', (order_id,))
    order = cursor.fetchone()

    if order:
        company = order[12]

        # Map the company to the corresponding vehicle table
        vehicle_table_map = {
            'MontasaHN': 'Vehicles',
            'MontasaCR': 'vehiculosCR',
            'Monhaco': 'vehiculosMonhaco'
        }
        vehicle_table = vehicle_table_map.get(company, 'Vehicles')

        # Fetch horometro and previously assigned supplies from the corresponding vehicle table
        cursor.execute(f'SELECT Horometro, InsumosMantenimiento, InsumosReparo FROM {vehicle_table} WHERE VehicleID = %s', (order[1],))
        vehicle_row = cursor.fetchone()
        horometro = vehicle_row[0] if vehicle_row else 0
        insumos_mantenimiento = vehicle_row[1] if vehicle_row and vehicle_row[1] else ''
        insumos_reparo = vehicle_row[2] if vehicle_row and vehicle_row[2] else ''

        # Fetch mechanics assigned to this work order
        cursor.execute('''
            SELECT m.MechanicID, m.Name
            FROM Mechanics m
            JOIN MechanicWorkOrder mwo ON m.MechanicID = mwo.MechanicID
            WHERE mwo.OrderID = %s
        ''', (order_id,))
        assigned_mechanics = cursor.fetchall()

        # Fetch all mechanics for assignment
        cursor.execute('SELECT MechanicID, Name FROM Mechanics')
        all_mechanics = cursor.fetchall()

        # Exclude mechanics already assigned to the work order
        assigned_mechanic_ids = {mechanic[0] for mechanic in assigned_mechanics}
        available_mechanics = [mechanic for mechanic in all_mechanics if mechanic[0] not in assigned_mechanic_ids]

        # Fetch assigned and available supplies
        cursor.execute('SELECT CodigoSap, Description, Quantity, Status FROM WorkOrderSupplies WHERE OrderID = %s', (order_id,))
        assigned_supplies = cursor.fetchall()

        cursor.execute('SELECT CodigoSap, Description FROM Supplies')
        available_supplies = cursor.fetchall()

        # Fetch previously assigned supplies for maintenance and repair
        previously_assigned_supplies = []
        if order[3] == 'Maintenance' and insumos_mantenimiento:
            insumos_ids = insumos_mantenimiento.split(',')
            format_strings = ','.join(['%s'] * len(insumos_ids))
            cursor.execute(f'SELECT CodigoSap, Description FROM Supplies WHERE CodigoSap IN ({format_strings})', tuple(insumos_ids))
            previously_assigned_supplies = cursor.fetchall()
        elif order[3] == 'Repair' and insumos_reparo:
            insumos_ids = insumos_reparo.split(',')
            format_strings = ','.join(['%s'] * len(insumos_ids))
            cursor.execute(f'SELECT CodigoSap, Description FROM Supplies WHERE CodigoSap IN ({format_strings})', tuple(insumos_ids))
            previously_assigned_supplies = cursor.fetchall()

        # Fetch waiting time start if currently waiting
        wait_time_start = None
        if order[13]:  # currently_waiting is True
            cursor.execute('SELECT StartTime FROM supplywaittimes WHERE OrderID = %s ORDER BY StartTime DESC LIMIT 1', (order_id,))
            wait_time_row = cursor.fetchone()
            if wait_time_row:
                wait_time_start = wait_time_row[0]

        cursor.close()
        conn.close()

        order_dict = {
            'order_id': order[0],
            'vehicle_id': order[1],
            'vehicle_name': order[2],
            'work_type': order[3],
            'description': order[4],
            'status': order[5],
            'Lugar': order[6],
            'Dueno': order[7],
            'Marca': order[8],
            'diagnostico': order[9],
            'trabajoRealizado': order[10],
            'created_time': order[11],
            'currently_waiting': order[13],
            'wait_time_start': wait_time_start,
            'horometro': horometro
        }

        return render_template('work_order_detail.html', 
                               order=order_dict, 
                               company=company, 
                               available_supplies=available_supplies, 
                               assigned_supplies=assigned_supplies, 
                               mechanics=assigned_mechanics, 
                               all_mechanics=available_mechanics,
                               previously_assigned_supplies=previously_assigned_supplies)
    else:
        cursor.close()
        conn.close()
        return 'Work Order not found', 404





#work order detail

#update work orders
@app.route('/update_work_order/<int:order_id>', methods=['POST'])
def update_work_order(order_id):
    descripcion = request.form.get('descripcion')
    diagnostico = request.form.get('diagnostico')
    trabajoRealizado = request.form.get('trabajoRealizado')



    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Update the work order with the new values
        update_query = '''
            UPDATE WorkOrders
            SET description = %s, Diagnostico = %s, TrabajoRealizado = %s
            WHERE OrderID = %s
        '''
        
        cursor.execute(update_query, (descripcion, diagnostico, trabajoRealizado, order_id))
        conn.commit()
        
        flash('Work order updated successfully.', 'success')
    except mysql.connector.Error as err:
        
        flash(f'Error updating work order: {err.msg}', 'error')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
        

    return redirect(url_for('work_order_detail', order_id=order_id))




#update work orders

#work order Records
@app.route('/work_orders_record/<company>', methods=['GET'])
def work_orders_record(company):
    order_id_query = request.args.get('order_id', '')
    description_query = request.args.get('description', '')
    vehicle_name_query = request.args.get('vehicle_name', '')
    start_date_query = request.args.get('start_date', '')
    end_date_query = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))

    conn = get_db_connection()
    cursor = conn.cursor()

    # Map the company to the corresponding vehicle table
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Obtain order IDs and vehicle names for filter options, filtering by company
    cursor.execute(f'''
        SELECT DISTINCT OrderID
        FROM WorkOrders
        WHERE Status = "Completed" AND Empresa = %s
        ORDER BY OrderID DESC
    ''', (company,))
    order_ids = [row[0] for row in cursor.fetchall()]

    cursor.execute(f'''
        SELECT DISTINCT IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE wo.Empresa = %s
        ORDER BY VehicleName
    ''', (company,))
    vehicle_names = [row[0] for row in cursor.fetchall()]

    # Get current page and set the number of items per page
    page = int(request.args.get('page', 1))
    items_per_page = 15
    offset = (page - 1) * items_per_page

    # Construct the SQL query based on search parameters, filtering by company
    query_conditions = ["wo.Status = 'Completed'", "wo.Empresa = %s"]
    query_params = [company]

    if order_id_query:
        query_conditions.append("wo.OrderID = %s")
        query_params.append(order_id_query)

    if description_query:
        query_conditions.append("wo.Description LIKE %s")
        query_params.append(f"%{description_query}%")

    if vehicle_name_query:
        query_conditions.append(f"(v.VehicleName = %s OR vc.VehicleName = %s)")
        query_params.extend([vehicle_name_query, vehicle_name_query])

    if start_date_query and end_date_query:
        query_conditions.append("wo.FinishedTime BETWEEN %s AND %s")
        query_params.extend([start_date_query, end_date_query + " 23:59:59"])

    where_clause = " AND ".join(query_conditions)
    sql_query = f'''
        SELECT wo.OrderID, wo.VehicleID, IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName, wo.WorkType, wo.Description, wo.CreatedTime, wo.FinishedTime
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE {where_clause}
        ORDER BY wo.OrderID DESC
        LIMIT %s OFFSET %s
    '''
    query_params.extend([items_per_page, offset])
    cursor.execute(sql_query, tuple(query_params))
    work_orders = cursor.fetchall()

    # Get the total number of work orders
    cursor.execute(f'''
        SELECT COUNT(*) as total
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE {where_clause}
    ''', tuple(query_params[:-2]))  # Exclude the LIMIT and OFFSET params
    total_orders = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    total_pages = (total_orders + items_per_page - 1) // items_per_page

    return render_template('work_orders_record.html', 
                           work_orders=work_orders, 
                           order_ids=order_ids, 
                           vehicle_names=vehicle_names, 
                           selected_order_id=order_id_query, 
                           selected_vehicle_name=vehicle_name_query, 
                           selected_description=description_query, 
                           start_date_query=start_date_query, 
                           end_date_query=end_date_query, 
                           company=company,
                           page=page,
                           total_pages=total_pages)


#work order Records
from datetime import datetime, timedelta
@app.route('/work_order_record_detail/<company>/<int:order_id>')
def work_order_record_detail(company, order_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Map the company to the corresponding vehicle table
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Fetch the details of the completed work order
    cursor.execute(f'''
        SELECT wo.OrderID, wo.VehicleID, IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName, wo.WorkType,
               wo.Description, wo.CreatedTime, wo.FinishedTime, wo.Lugar, wo.Dueno, wo.Marca,
               wo.Diagnostico, wo.TrabajoRealizado, wo.WorkedTime
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE wo.OrderID = %s AND wo.Status = 'Completed' AND wo.Empresa = %s
    ''', (order_id, company))
    order = cursor.fetchone()

    # Fetch all mechanics and their work times assigned to this work order
    cursor.execute('''
        SELECT m.MechanicID, m.Name,
            TIMESTAMPDIFF(SECOND, t.StartTime, t.EndTime) AS MechanicTime, t.StartTime, t.EndTime
        FROM Mechanics m
        JOIN TimeTracking t ON m.MechanicID = t.MechanicID
        WHERE t.OrderID = %s
    ''', (order_id,))
    mechanics = cursor.fetchall()

    # Fetch all supplies assigned to this work order
    cursor.execute('''
        SELECT s.SupplyID, s.Description, ws.Quantity, ws.ReceivedDate
        FROM Supplies s
        JOIN WorkOrderSupplies ws ON s.SupplyID = ws.SupplyID
        WHERE ws.OrderID = %s AND ws.Status = 'Recibido'
    ''', (order_id,))
    supplies = cursor.fetchall()

    cursor.close()
    conn.close()

    if order:
        # Convert the order tuple to a dictionary for easier access in the template
        order_dict = {
            'order_id': order[0],
            'vehicle_id': order[1],
            'vehicle_name': order[2],
            'work_type': order[3],
            'description': order[4],
            'start_date': order[5],
            'end_date': order[6],
            'Lugar': order[7],
            'Dueno': order[8],
            'Marca': order[9],
            'diagnostico': order[10],
            'trabajorealizado': order[11],
            'worked_time': order[12]
        }

        # Prepare mechanic details
        mechanic_details = []
        for mechanic in mechanics:
            mechanic_id = mechanic[0]
            mechanic_name = mechanic[1]
            mechanic_time_seconds = float(mechanic[2])
            mechanic_time = str(timedelta(seconds=mechanic_time_seconds))
            mechanic_details.append({
                'id': mechanic_id,
                'name': mechanic_name,
                'time': mechanic_time,
                'start_time': mechanic[3],
                'end_time': mechanic[4]
            })

        # Prepare supply details
        supply_details = []
        for supply in supplies:
            supply_details.append({
                'id': supply[0],
                'description': supply[1],
                'quantity': supply[2],
                'received_date': supply[3]
            })

        return render_template('work_order_record_detail.html',
                               order=order_dict,
                               company=company,
                               mechanics=mechanic_details,
                               supplies=supply_details,
                               timedelta=timedelta)
    else:
        return 'Work Order not found', 404






#work order record detail





#mechanic quality check

@app.route('/create_mechanic_quality_check/<int:work_order_id>', methods=['GET'])
def create_mechanic_quality_check(conn, work_order_id):
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO MechanicQualityChecks (WorkOrderID)
        VALUES (%s)
    ''', (work_order_id,))
    conn.commit()
    cursor.close()


@app.route('/mechanic_quality_check/<int:order_id>')
def mechanic_quality_check(order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get the company and client SAP code from the work order
    cursor.execute('''
        SELECT Empresa, ClienteCodigoSap FROM WorkOrders WHERE OrderID = %s
    ''', (order_id,))
    work_order = cursor.fetchone()
    company = work_order['Empresa']
    cliente_codigo_sap = work_order['ClienteCodigoSap']

    # Map the company to the corresponding vehicle table
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }

    if cliente_codigo_sap:
        vehicle_table = 'VehiclesClientes'
    else:
        vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Fetch checklist, mechanic's name, and boss's name based on WorkOrderID and company
    cursor.execute(f'''
        SELECT mq.*, m.Name AS MechanicName, b.Nombre AS BossName, b.BossID, v.VehicleName, w.Empresa
        FROM MechanicQualityChecks mq
        LEFT JOIN Mechanics m ON mq.MechanicID = m.MechanicID
        LEFT JOIN Boss b ON mq.BossID = b.BossID
        LEFT JOIN WorkOrders w ON mq.WorkOrderID = w.OrderID
        LEFT JOIN {vehicle_table} v ON w.VehicleID = v.VehicleID
        WHERE mq.WorkOrderID = %s
    ''', (order_id,))
    checklist = cursor.fetchone()

    cursor.close()
    conn.close()

    if checklist:
        return render_template('mechanic_quality_check.html', checklist=checklist, order_id=order_id, vehicle_name=checklist['VehicleName'])
    else:
        flash('No checklist found for this work order.', 'error')
        return redirect(url_for('work_order_detail', order_id=order_id, company=company))







@app.route('/update_mechanic_quality_checklist/<int:order_id>', methods=['POST'])
def update_mechanic_quality_checklist(order_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Fields to update, including 'MechanicID' and 'BossID'
        fields = [
            'BossID', 'MechanicID', 'CambioAceiteMotor', 'CambioFiltroMotor', 'CambioFiltroAire', 'CambioFiltroCombustible',
            'RevisionCambioBandaMotor', 'CambioRefrigeranteRadiador', 'CambioTapaderaDistribuidorRotor',
            'CambioCandelas', 'CambioCableCandela', 'CambioDeAceiteHidraulico', 'CambioDeFiltroHidraulico',
            'LimpiezaDeFiltroDeMalla', 'CambioDeEmpaqueDeTapadera', 'RevisarYPruebaDeSistemaHidraulico',
            'RevisionGeneralDeManguerasHidraulicas', 'RevisionSoporteMotor', 'RevisionInyectores',
            'RevisionTiempoMotor', 'RevisionTermostato', 'RevisionMultipleAdmision', 'RevisionTornillosManifold',
            'RevisionRPM', 'RevisionCargaAlternador', 'CambioAceiteTransmision', 'CambioFiltroTransmision',
            'CambioAceiteDiferencial', 'RevisionCambioManguera', 'RevisionFugas', 'RevisionBombasFreno',
            'RevisionFricciones', 'RevisionBalinerasRetenedores', 'RevisionEjeDireccion', 'AjustesPedalesFreno',
            'RevisarLubricarEjesDelanteros', 'RevisionTambor', 'RevisarFuncionamientoIndicadores', 'RevisarMotorArranque',
            'RevisarArnesEquipo', 'RevisarSistemaCarga', 'SocarTerminalBateria', 'RevisionAmpBateria',
            'RevisionTaponColadorPolvo', 'RevisionSistemaLucesDelanteras', 'RevisionSistemaLucesTraseras',
            'LucesStop', 'ViasDireccionales', 'LucesRetroceso', 'AlarmaRetroceso', 'LuzEstroboscopico', 'Claxon',
            'CinturonSeguridad', 'Retrovisores', 'Extintor', 'RevisarTensionCadena', 'RevisarAjustesEmergencias',
            'LavadoGeneral', 'LubricacionEngraseTorreCadena', 'CalibracionValvulaMotor', 'AdditionalNotes'
        ]

        # Collect values for each field
        update_values = []
        for field in fields:
            value = request.form.get(field)
            if field in ['MechanicID', 'BossID']:
                update_values.append(value if value else None)
            elif field == 'AdditionalNotes':
                update_values.append(value)
            else:
                if value == 'si':
                    update_values.append(1)
                elif value == 'no':
                    update_values.append(0)
                elif value == 'na':
                    update_values.append(None)  # Assuming 'na' should be stored as NULL
                else:
                    update_values.append(value)
        update_values.append(order_id)  # Append the WorkOrderID at the end

        # Generating the SQL Update Statement
        update_query = f"UPDATE MechanicQualityChecks SET {', '.join(f'{field} = %s' for field in fields)} WHERE WorkOrderID = %s"



        cursor.execute(update_query, update_values)
        conn.commit()
        flash('Checklist updated successfully.', 'success')
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'Database error: {err}', 'error')
        print("SQL Error: ", err)
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_detail', order_id=order_id))

@app.route('/mechanic_quality_check_record/<int:order_id>', methods=['GET'])
def mechanic_quality_check_record(order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get the company and client SAP code from the work order
    cursor.execute('''
        SELECT Empresa, ClienteCodigoSap FROM WorkOrders WHERE OrderID = %s
    ''', (order_id,))
    work_order = cursor.fetchone()
    company = work_order['Empresa']
    cliente_codigo_sap = work_order['ClienteCodigoSap']

    # Map the company to the corresponding vehicle table
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }

    if cliente_codigo_sap:
        vehicle_table = 'VehiclesClientes'
    else:
        vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Fetch the mechanic quality check record
    cursor.execute(f'''
        SELECT mq.*, m.Name AS MechanicName, b.Nombre AS BossName, b.BossID, v.VehicleName, w.Empresa
        FROM MechanicQualityChecks mq
        LEFT JOIN Mechanics m ON mq.MechanicID = m.MechanicID
        LEFT JOIN Boss b ON mq.BossID = b.BossID
        LEFT JOIN WorkOrders w ON mq.WorkOrderID = w.OrderID
        LEFT JOIN {vehicle_table} v ON w.VehicleID = v.VehicleID
        WHERE mq.WorkOrderID = %s
    ''', (order_id,))
    checklist = cursor.fetchone()

    cursor.close()
    conn.close()

    if checklist:
        return render_template('mechanic_quality_check_record.html', checklist=checklist, order_id=order_id, vehicle_name=checklist['VehicleName'], company=company)
    else:
        flash('No checklist found for this work order.', 'error')
        return redirect(url_for('work_order_detail', order_id=order_id, company=company))









#mechanic quality check

#logistics Quality check

def create_logistics_quality_check(conn, order_id):
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO LogisticsQualityChecklist (OrderID)
        VALUES (%s)
    ''', (order_id,))
    conn.commit()
    cursor.close()

@app.route('/logistics_quality_check/<int:order_id>')
def logistics_quality_check(order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch checklist and operator's and boss's name based on OrderID
    cursor.execute('''
        SELECT lqc.*, o.NombreOperador AS OperatorName, o.IdOperador, b.Nombre AS BossName, b.BossID
        FROM LogisticsQualityChecklist lqc
        LEFT JOIN Boss b ON lqc.BossID = b.BossID
        LEFT JOIN Operadores o ON lqc.IdOperador = o.IdOperador
        WHERE lqc.OrderID = %s
    ''', (order_id,))
    checklist = cursor.fetchone()

    cursor.close()
    conn.close()

    if checklist:
        return render_template('logistics_quality_check.html', checklist=checklist, order_id=order_id)
    else:
        flash('No checklist found for this work order.', 'error')
        return redirect(url_for('work_order_detail', order_id=order_id))



@app.route('/update_logistics_quality_checklist/<int:order_id>', methods=['POST'])
def update_logistics_quality_checklist(order_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Fields to update
        fields = [
            'BossID', 'IdOperador', 'AlarmaDePrecaucion', 'BotonDeEmergencia', 'LuzEstroboscopica', 'LucesDeTrabajo',
            'LucesDeStop', 'LucesDeVias', 'Baterias', 'Pito', 'TaponDeCombustible', 'LubricacionYEngrase',
            'SistemaHidraulico','BandasDeMotor', 'FugasDeAceites','TorreCompleta','LavadoGeneral','FrenosGeneral',
            'CilindroDeGas', 'Mangueras', 'Cuchillas', 'Balineras', 'Palancas', 'Pistones',
            'Cadenas', 'Shifter', 'GolpesOAbolladuras', 'IndicadoresTablero', 'CamaraFrontal', 'CamaraTrasera',
            'PantallaVisual', 'PinturaGeneral', 'Cinturones', 'Asientos', 'Extintor', 'ExtensionesHorquilla',
            'LlavesDeEncendido', 'KitAntiderrame', 'Conos', 'DelanteraIzquierda', 'DelanteraDerecha', 'TraseraIzquierda',
            'TraseraDerecha', 'RefrigeranteCoolant', 'LiquidoDeFrenos', 'AceiteHidraulico', 'AceiteMotor',
            'Combustible', 'ConectoresDeCorriente', 'CableDeEmergencia', 'Cargador', 'Botones', 'Canasta', 'Joystick',
            'Switch', 'AdditionalNotes'
        ]

        # Collect values for each field
        update_values = []
        for field in fields:
            value = request.form.get(field)
            if field in ['IdOperador', 'BossID']:
                update_values.append(value if value else None)
            elif field == 'AdditionalNotes':
                update_values.append(value)
            else:
                if value == 'si':
                    update_values.append(1)
                elif value == 'no':
                    update_values.append(0)
                elif value == 'na':
                    update_values.append(None)
                else:
                    update_values.append(value)

        update_values.append(order_id)  # Append the OrderID at the end

        # Construct the SQL update statement
        update_query = f"UPDATE LogisticsQualityChecklist SET {', '.join(f'{field} = %s' for field in fields)} WHERE OrderID = %s"



        cursor.execute(update_query, update_values)
        conn.commit()
        flash('Logistics quality checklist updated successfully.', 'success')
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'Database error: {err}', 'error')
        print("SQL Error: ", err)
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_detail', order_id=order_id))



@app.route('/update_logistics_quality_checklist_record/<company>/<int:order_id>', methods=['POST'])
def update_logistics_quality_checklist_record(company, order_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Fields to update
        fields = [
            'BossID', 'IdOperador', 'AlarmaDePrecaucion', 'BotonDeEmergencia', 'LuzEstroboscopica', 'LucesDeTrabajo',
            'LucesDeStop', 'LucesDeVias', 'Baterias', 'Pito', 'TaponDeCombustible', 'LubricacionYEngrase',
            'SistemaHidraulico','BandasDeMotor', 'FugasDeAceites','TorreCompleta','LavadoGeneral','FrenosGeneral',
            'CilindroDeGas', 'Mangueras', 'Cuchillas', 'Balineras', 'Palancas', 'Pistones',
            'Cadenas', 'Shifter', 'GolpesOAbolladuras', 'IndicadoresTablero', 'CamaraFrontal', 'CamaraTrasera',
            'PantallaVisual', 'PinturaGeneral', 'Cinturones', 'Asientos', 'Extintor', 'ExtensionesHorquilla',
            'LlavesDeEncendido', 'KitAntiderrame', 'Conos', 'DelanteraIzquierda', 'DelanteraDerecha', 'TraseraIzquierda',
            'TraseraDerecha', 'RefrigeranteCoolant', 'LiquidoDeFrenos', 'AceiteHidraulico', 'AceiteMotor',
            'Combustible', 'ConectoresDeCorriente', 'CableDeEmergencia', 'Cargador', 'Botones', 'Canasta', 'Joystick',
            'Switch', 'AdditionalNotes'
        ]

        # Collect values for each field
        update_values = []
        for field in fields:
            value = request.form.get(field)
            if field in ['IdOperador', 'BossID']:
                update_values.append(value if value else None)
            elif field == 'AdditionalNotes':
                update_values.append(value)
            else:
                if value == 'si':
                    update_values.append(1)
                elif value == 'no':
                    update_values.append(0)
                elif value == 'na':
                    update_values.append(None)
                else:
                    update_values.append(value)

        update_values.append(order_id)  # Append the OrderID at the end

        # Construct the SQL update statement
        update_query = f"UPDATE LogisticsQualityChecklist SET {', '.join(f'{field} = %s' for field in fields)} WHERE OrderID = %s"

        cursor.execute(update_query, update_values)
        conn.commit()
        flash('Logistics quality checklist updated successfully.', 'success')
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'Database error: {err}', 'error')
        print("SQL Error: ", err)
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_record_detail', company=company, order_id=order_id))



@app.route('/logistics_quality_check_record/<int:order_id>', methods=['GET'])
def logistics_quality_check_record(order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute('''
        SELECT lqc.*, o.NombreOperador AS OperatorName, o.IdOperador, b.Nombre AS BossName, b.BossID, w.Empresa, v.VehicleName
        FROM LogisticsQualityChecklist lqc
        LEFT JOIN Boss b ON lqc.BossID = b.BossID
        LEFT JOIN Operadores o ON lqc.IdOperador = o.IdOperador
        LEFT JOIN WorkOrders w ON lqc.OrderID = w.OrderID
        LEFT JOIN Vehicles v ON w.VehicleID = v.VehicleID AND w.Empresa = v.Empresa
        
        WHERE lqc.OrderID = %s
    ''', (order_id,))
    checklist = cursor.fetchone()

    cursor.close()
    conn.close()

    if checklist:
        return render_template('logistics_quality_check_record.html', checklist=checklist, order_id=order_id, company=checklist['Empresa'],vehicle_name=checklist['VehicleName'])
    else:
        flash('No checklist found for this work order.', 'error')
        return redirect(url_for('work_order_detail', order_id=order_id, company=checklist['Empresa']))



#logistics Quality check

#operator checklist
# Crear una entrada en operatorchecklist
def create_operator_checklist(conn, id_orden_salida):
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO OperatorChecklist (IdOrdenSalida)
        VALUES (%s)
    ''', (id_orden_salida,))
    conn.commit()
    cursor.close()


@app.route('/operator_quality_check/<int:order_id>')
def operator_quality_check(order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch checklist based on OrderID
    cursor.execute('''
        SELECT oc.*, o.NombreOperador AS OperatorName, o.IdOperador, b.Nombre AS BossName, b.BossID
        FROM OperatorChecklist oc
        LEFT JOIN Boss b ON oc.BossID = b.BossID
        LEFT JOIN Operadores o ON oc.IdOperador = o.IdOperador
        WHERE oc.IdOrdenSalida = %s
    ''', (order_id,))
    checklist = cursor.fetchone()

    # Fetch the company from the order details
    cursor.execute('SELECT Empresa FROM OrdenesdeSalida WHERE IdOrdenSalida = %s', (order_id,))
    order = cursor.fetchone()
    company = order['Empresa'] if order else None

    cursor.close()
    conn.close()

    if checklist:
        return render_template('operator_checklist.html', checklist=checklist, order_id=order_id, company=company)
    else:
        flash('No checklist found for this order.', 'error')
        return redirect(url_for('departure_order_detail', company=company, order_id=order_id))



@app.route('/update_operator_quality_checklist/<int:order_id>', methods=['POST'])
def update_operator_quality_checklist(order_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Fields to update
        fields = [
            'BossID', 'IdOperador', 'AlarmaDePrecaucion', 'BotonDeEmergencia', 'LuzEstroboscopica', 'LucesDeTrabajo',
            'LucesDeStop', 'LucesDeVias', 'Baterias', 'Pito', 'TaponDeCombustible', 'LubricacionYEngrase',
            'SistemaHidraulico','BandasDeMotor', 'FugasDeAceites','TorreCompleta','LavadoGeneral','FrenosGeneral',
            'CilindroDeGas', 'Mangueras', 'Cuchillas', 'Balineras', 'Palancas', 'Pistones',
            'Cadenas', 'Shifter', 'GolpesOAbolladuras', 'IndicadoresTablero', 'CamaraFrontal', 'CamaraTrasera',
            'PantallaVisual', 'PinturaGeneral', 'Cinturones', 'Asientos', 'Extintor', 'ExtensionesHorquilla',
            'LlavesDeEncendido', 'KitAntiderrame', 'Conos', 'DelanteraIzquierda', 'DelanteraDerecha', 'TraseraIzquierda',
            'TraseraDerecha', 'RefrigeranteCoolant', 'LiquidoDeFrenos', 'AceiteHidraulico', 'AceiteMotor',
            'Combustible', 'ConectoresDeCorriente', 'CableDeEmergencia', 'Cargador', 'Botones', 'Canasta', 'Joystick',
            'Switch', 'AdditionalNotes'
        ]

        # Collect values for each field
        update_values = []
        for field in fields:
            value = request.form.get(field)
            if field in ['IdOperador', 'BossID']:
                update_values.append(value if value else None)
            elif field == 'AdditionalNotes':
                update_values.append(value)
            else:
                if value == 'si':
                    update_values.append(1)
                elif value == 'no':
                    update_values.append(0)
                elif value == 'na':
                    update_values.append(None)
                else:
                    update_values.append(value)

        update_values.append(order_id)  # Append the OrderID at the end

        # Construct the SQL update statement
        update_query = f"UPDATE OperatorChecklist SET {', '.join(f'{field} = %s' for field in fields)} WHERE IdOrdenSalida = %s"

        cursor.execute(update_query, update_values)
        conn.commit()
        flash('Operator quality checklist updated successfully.', 'success')
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'Database error: {err}', 'error')
        print("SQL Error: ", err)
    finally:
        cursor.close()
        conn.close()

    # Ensure you pass both company and order_id
    company = request.form.get('company')  # Assuming company is passed in the form data
    return redirect(url_for('departure_order_detail', company=company, order_id=order_id))












# Complete Work Order
@app.route('/complete_work_order', methods=['POST'])
def complete_work_order():
    order_id = request.form['order_id']
    finished_time = datetime.now()  # Capture the current time as finished time

    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Complete the work order and set the finished time
        cursor.execute('UPDATE WorkOrders SET Status = "Completed", FinishedTime = %s WHERE OrderID = %s', (finished_time, order_id))

        cursor.execute('SELECT Empresa, ClienteCodigoSap FROM WorkOrders WHERE OrderID = %s', (order_id,))
        order_info = cursor.fetchone()
        company = order_info[0]
        cliente_codigo_sap = order_info[1]

        # Determine the appropriate vehicle table
        vehicle_table_map = {
            'MontasaHN': 'Vehicles',
            'MontasaCR': 'vehiculosCR',
            'Monhaco': 'vehiculosMonhaco'
        }

        if cliente_codigo_sap:
            vehicle_table = 'VehiclesClientes'
        else:
            vehicle_table = vehicle_table_map.get(company, 'Vehicles')

        cursor.execute('SELECT VehicleID FROM WorkOrders WHERE OrderID = %s', (order_id,))
        vehicle_info = cursor.fetchone()
        if not vehicle_info:
            raise ValueError(f"No vehicle found for order ID {order_id}")
        vehicle_id = vehicle_info[0]

        cursor.execute(f'SELECT Status FROM {vehicle_table} WHERE VehicleID = %s', (vehicle_id,))
        vehicle_status_info = cursor.fetchone()
        if not vehicle_status_info:
            raise ValueError(f"No status found for vehicle ID {vehicle_id} in table {vehicle_table}")
        current_status = vehicle_status_info[0]

        # Update vehicle status based on the current status
        if current_status == 'En Taller':
            new_status = 'Disponible'
        elif current_status == 'Reparación Externa':
            new_status = 'En Renta'
        else:
            new_status = current_status  # Keep the current status if it doesn't match the specified conditions

        cursor.execute(f'UPDATE {vehicle_table} SET Status = %s WHERE VehicleID = %s', (new_status, vehicle_id))

        # Check for assigned mechanics in MechanicWorkOrder table
        cursor.execute('SELECT MechanicID FROM MechanicWorkOrder WHERE OrderID = %s', (order_id,))
        assigned_mechanics = cursor.fetchall()
        for mechanic in assigned_mechanics:
            mechanic_id = mechanic[0]
            unassign_mechanic_logic(order_id, mechanic_id)

        conn.commit()
    except mysql.connector.Error as err:
        print(f"SQL Error: {err}")
        conn.rollback()  # Roll back in case of error
    except ValueError as err:
        print(f"Value Error: {err}")
        conn.rollback()  # Roll back in case of error
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_orders', company=company))





#complete work order

#start Supply wait

@app.route('/start_supply_wait/<int:order_id>', methods=['POST'])
def start_supply_wait(order_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Check if there's an ongoing supply wait for this order
        cursor.execute('SELECT * FROM SupplyWaitTimes WHERE OrderID = %s AND EndTime IS NULL', (order_id,))
        active_wait = cursor.fetchone()

        if not active_wait:
            # No active wait, start a new wait period
            cursor.execute('INSERT INTO SupplyWaitTimes (OrderID, StartTime) VALUES (%s, NOW())', (order_id,))
            cursor.execute('UPDATE WorkOrders SET Status = %s, currently_waiting = 1 WHERE OrderID = %s', ('SupplyWait', order_id))
            conn.commit()
            flash('Supply wait time started.', 'success')  # Using flash for messaging
        else:
            # An active wait exists
            flash('Supply wait time is already active.', 'info')  # Using flash for messaging

    except mysql.connector.Error as err:
        print("SQL Error: ", err)
        flash('Error starting supply wait time.', 'error')  # Using flash for messaging
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_detail', order_id=order_id))

@app.route('/stop_supply_wait/<int:order_id>', methods=['POST'])
def stop_supply_wait(order_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Update the SupplyWaitTimes table to set the end time
        cursor.execute('UPDATE SupplyWaitTimes SET EndTime = NOW() WHERE OrderID = %s AND EndTime IS NULL', (order_id,))
        # Update currently_waiting field to 0
        cursor.execute('UPDATE WorkOrders SET currently_waiting = 0 WHERE OrderID = %s', (order_id,))

        conn.commit()
        flash('Supply wait time stopped.', 'success')
    except mysql.connector.Error as err:
        print("SQL Error: ", err)
        flash('Database error occurred', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('work_order_detail', order_id=order_id))


#stop supply wait

#mechanic work hours

@app.route('/mechanic_work_hours', methods=['GET', 'POST'])
def mechanic_work_hours():
    filter_mechanic_id = request.args.get('filter_mechanic_id')
    search_query = request.args.get('search_query', '')

    conn = get_db_connection()
    mechanics = []
    work_hours = []

    try:
        cursor = conn.cursor()

        # Fetch list of mechanics for the dropdown
        cursor.execute('SELECT MechanicID, Name FROM Mechanics ORDER BY Name')
        mechanics = cursor.fetchall()

        # Build the query based on filters
        query = '''
            SELECT m.Name, tt.OrderID, tt.StartTime, tt.EndTime, TIMEDIFF(tt.EndTime, tt.StartTime) AS Duration
            FROM TimeTracking tt
            JOIN Mechanics m ON tt.MechanicID = m.MechanicID
        '''
        conditions = []
        params = []

        if filter_mechanic_id:
            conditions.append('m.MechanicID = %s')
            params.append(filter_mechanic_id)

        if search_query:
            conditions.append('tt.OrderID LIKE %s')
            params.append('%' + search_query + '%')

        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)

        query += ' ORDER BY tt.StartTime DESC'
        cursor.execute(query, params)
        work_hours = cursor.fetchall()

    finally:
        cursor.close()
        conn.close()

    return render_template('mechanic_work_hours.html', work_hours=work_hours, mechanics=mechanics)


#mechanic work hours

#Ordenes de SERVICIO fin


#ordenes de SALIDA Principio

@app.route('/add_departure_order/<company>', methods=['GET', 'POST'])
def add_departure_order(company):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Refresh clients from SAP (No se imprime aquí porque solicitaste no hacerlo)
    refresh_clients_from_sap()

    # Mapeo de la empresa a su tabla de vehículos correspondiente
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }

    # Obtén la tabla de vehículos correcta según la empresa
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')  # 'Vehicles' como predeterminado si no se encuentra la empresa

    print(f'Debug: Usando la tabla de vehículos {vehicle_table} para la empresa {company}')

    # Busca vehículos disponibles en la tabla correspondiente
    cursor.execute(f'SELECT VehicleID, VehicleName, Horometro FROM {vehicle_table} WHERE Status = %s', ('Disponible',))
    available_vehicles = cursor.fetchall()
    

    # Consulta operadores
    cursor.execute('SELECT IdOperador, NombreOperador FROM Operadores')
    operadores = cursor.fetchall()
    print(f'Debug: Operadores consultados: {operadores}')

    # Consulta clientes
    cursor.execute('SELECT ClienteID, Nombre FROM Clientes')
    clientes = cursor.fetchall()
    

    if request.method == 'POST':
        vehicle_id = request.form['vehicle_id']
        operator_id = request.form['operator_id']
        location = request.form['location']
        comments = request.form['comments']
        orden_de_sap = request.form['orden_de_sap']
        cliente_id = request.form['cliente_id']

        # Verifica el estado actual del vehículo en la tabla correspondiente
        cursor.execute(f'SELECT Status, Horometro, VehicleName FROM {vehicle_table} WHERE VehicleID = %s', (vehicle_id,))
        vehicle_status = cursor.fetchone()
        print(f'Debug: Estado del vehículo {vehicle_id} consultado: {vehicle_status}')

        if vehicle_status and vehicle_status[0] == 'En Taller':
            # Si el vehículo está en taller, muestra un mensaje de error
            flash('El vehículo seleccionado está en taller y no puede ser asignado a una orden de salida.', 'error')
        else:
            try:
                horometro_salida = vehicle_status[1]  # Obtener el horómetro actual del vehículo
                vehicle_name = vehicle_status[2]  # Obtener el nombre del vehículo

                # Inserta en OrdenesdeSalida
                insert_query = f'''
                    INSERT INTO OrdenesdeSalida (VehicleID, VehicleName, IdOperador, NombreOperador, Comentarios, Ubicacion, Empresa, HoraCreado, OrdenDeSap, ClienteID, HorometroSalida)
                    VALUES (%s, %s, %s, (SELECT NombreOperador FROM Operadores WHERE IdOperador = %s), %s, %s, %s, NOW(), %s, %s, %s)
                '''
                cursor.execute(insert_query, (vehicle_id, vehicle_name, operator_id, operator_id, comments, location, company, orden_de_sap, cliente_id, horometro_salida))
                print(f'Debug: Orden de salida insertada con VehicleID {vehicle_id} y Operador {operator_id}')

                cursor.execute('''
                    INSERT INTO OperatorChecklist (IdOrdenSalida)
                    VALUES (%s)
                ''', (cursor.lastrowid,))
                print('Debug: Checklist de operador insertado')

                # Verifica si el nombre del vehículo no es genérico antes de actualizar el estado
                if vehicle_name.lower() not in ["generico", "genérico"]:
                    # Actualiza el estado y la ubicación del vehículo en la tabla correspondiente
                    update_vehicle_query = f'''
                        UPDATE {vehicle_table}
                        SET Status = 'En Renta', Ubicacion = %s
                        WHERE VehicleID = %s
                    '''
                    cursor.execute(update_vehicle_query, (location, vehicle_id))
                    print(f'Debug: Estado del vehículo {vehicle_id} actualizado a "En Renta" con ubicación {location}')

                conn.commit()
                flash('Orden de salida agregada, ubicación y estado del vehículo actualizados.', 'success')
            except mysql.connector.Error as err:
                conn.rollback()
                print(f'Debug: Error al agregar la orden de salida: {err}')
                flash(f'Error al agregar la orden de salida: {err}', 'error')
            finally:
                cursor.close()
                conn.close()

        return redirect(url_for('active_departure_orders', company=company))

    cursor.close()
    conn.close()

    return render_template('CrearOrdenSalida.html', available_vehicles=available_vehicles, operadores=operadores, clientes=clientes, company=company)





vehicle_table_map = {
    'MontasaHN': 'Vehicles',
    'MontasaCR': 'vehiculosCR',
    'Monhaco': 'vehiculosMonhaco'
}

@app.route('/active_departure_orders/<company>')
def active_departure_orders(company):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Mapeo de la compañía a la tabla de vehículos correspondiente
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')  # Usa 'Vehicles' como predeterminado si la compañía no se encuentra

    # Obtener los nombres de los vehículos para el dropdown, filtrando por empresa
    cursor.execute(f'SELECT DISTINCT VehicleName FROM {vehicle_table} WHERE Empresa = %s ORDER BY VehicleID', (company,))
    vehicle_names = cursor.fetchall()

    # Obtener los nombres de los clientes con órdenes de salida activas
    cursor.execute(f'''
        SELECT DISTINCT c.ClienteID, c.Nombre
        FROM OrdenesdeSalida os
        JOIN Clientes c ON os.ClienteID = c.ClienteID
        WHERE os.HoraRegreso IS NULL AND os.Empresa = %s
    ''', (company,))
    client_names = cursor.fetchall()

    # Recuperar el vehículo seleccionado, el cliente seleccionado y las fechas desde la solicitud
    selected_vehicle_name = request.args.get('vehicle_name', '')
    selected_client_id = request.args.get('client_id', '')
    start_date = request.args.get('start_date', '2024-01-01')  # Fecha de inicio por defecto
    end_date = request.args.get('end_date', (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d'))  # Fecha de fin al día siguiente
    page = int(request.args.get('page', 1))
    per_page = 20

    # Construir la consulta SQL con los filtros aplicados, incluyendo la empresa
    query = f'''
        SELECT os.IdOrdenSalida, os.VehicleID, v.VehicleName, os.IdOperador, op.NombreOperador as NombreOperador, os.HoraCreado, os.Comentarios, os.OrdenDeSAP, os.Ubicacion, c.Nombre as ClienteNombre
        FROM OrdenesdeSalida os
        JOIN {vehicle_table} v ON os.VehicleID = v.VehicleID
        JOIN Operadores op ON os.IdOperador = op.IdOperador
        JOIN Clientes c ON os.ClienteID = c.ClienteID
        WHERE os.HoraRegreso IS NULL AND os.Empresa = %s
    '''
    params = [company]

    if selected_vehicle_name:
        query += ' AND v.VehicleName = %s'
        params.append(selected_vehicle_name)

    if selected_client_id:
        query += ' AND c.ClienteID = %s'
        params.append(selected_client_id)

    # Añadir el filtro de rango de fechas a la consulta
    query += ' AND os.HoraCreado BETWEEN %s AND %s'
    params.extend([start_date, end_date])

    # Contar el número total de resultados
    count_query = f"SELECT COUNT(*) as total FROM ({query}) as subquery"
    cursor.execute(count_query, tuple(params))
    total_results = cursor.fetchone()['total']

    total_pages = (total_results // per_page) + (1 if total_results % per_page > 0 else 0)
    offset = (page - 1) * per_page

    # Añadir límite y desplazamiento para la paginación
    query += f' ORDER BY os.IdOrdenSalida DESC LIMIT {per_page} OFFSET {offset}'

    cursor.execute(query, tuple(params))
    active_departure_orders = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('active_departure_orders.html',
                           active_departure_orders=active_departure_orders,
                           vehicle_names=vehicle_names,
                           client_names=client_names,
                           selected_vehicle_name=selected_vehicle_name,
                           selected_client_id=selected_client_id,
                           start_date=start_date,
                           end_date=end_date,
                           company=company,
                           page=page,
                           total_pages=total_pages)





@app.route('/departure_order_detail/<company>/<int:order_id>', methods=['GET', 'POST'])
def departure_order_detail(company, order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    if request.method == 'POST':
        # Recuperar los datos del formulario
        comentarios = request.form.get('comentarios')
        operador_nombre = request.form.get('operador')
        nueva_ubicacion = request.form.get('ubicacion')

        # Obtener el IdOperador basado en el nombre del operador
        cursor.execute('SELECT IdOperador FROM Operadores WHERE NombreOperador = %s', (operador_nombre,))
        operador = cursor.fetchone()

        if operador:
            id_operador = operador['IdOperador']

            try:
                # Actualizar los detalles de la orden de salida
                cursor.execute('''
                    UPDATE OrdenesdeSalida
                    SET Comentarios = %s, IdOperador = %s, Ubicacion = %s
                    WHERE IdOrdenSalida = %s AND Empresa = %s
                ''', (comentarios, id_operador, nueva_ubicacion, order_id, company))
                conn.commit()

                flash('Orden de salida actualizada exitosamente.', 'success')
            except mysql.connector.Error as err:
                conn.rollback()
                flash(f'Error al actualizar la orden de salida: {err}', 'error')

    # Obtener detalles de la orden de salida
    cursor.execute(f'''
        SELECT os.IdOrdenSalida, os.VehicleID, v.VehicleName, os.IdOperador, o.NombreOperador,
               os.HoraCreado, os.Comentarios, os.HorometroSalida, os.HorometroRegreso,
               os.HoraRegreso, os.Ubicacion, c.Nombre as ClienteNombre
        FROM OrdenesdeSalida os
        JOIN {vehicle_table} v ON os.VehicleID = v.VehicleID
        JOIN Operadores o ON os.IdOperador = o.IdOperador
        JOIN Clientes c ON os.ClienteID = c.ClienteID
        WHERE os.IdOrdenSalida = %s AND os.Empresa = %s
    ''', (order_id, company))
    order_detail = cursor.fetchone()

    if not order_detail:
        cursor.close()
        conn.close()
        return 'Departure Order not found', 404

    # Obtener la lista de operadores
    cursor.execute('SELECT NombreOperador FROM Operadores')
    operadores = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('departure_order_detail.html', order=order_detail, operadores=operadores, company=company, order_id=order_id) 


@app.route('/update_departure_order/<company>/<int:order_id>', methods=['POST'])
def update_departure_order(company, order_id):
    # Recuperar los datos del formulario
    horometroSalida = request.form.get('horometroSalida')
    horometroRegreso = request.form.get('horometroRegreso')
    comentarios = request.form.get('comentarios')
    operador_nombre = request.form.get('operador')
    nueva_ubicacion = request.form.get('ubicacion')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    try:
        # Obtener el IdOperador basado en el nombre del operador
        cursor.execute('SELECT IdOperador FROM Operadores WHERE NombreOperador = %s', (operador_nombre,))
        operador = cursor.fetchone()

        if operador:
            id_operador = operador['IdOperador']

            # Actualizar los detalles de la orden de salida
            cursor.execute('''
                UPDATE OrdenesdeSalida
                SET HorometroSalida = %s, HorometroRegreso = %s, Comentarios = %s, IdOperador = %s, Ubicacion = %s
                WHERE IdOrdenSalida = %s AND Empresa = %s
            ''', (horometroSalida, horometroRegreso, comentarios, id_operador, nueva_ubicacion, order_id, company))

            # Actualizar el horómetro actual del vehículo, si se proporciona el horómetro de regreso
            if horometroRegreso:
                cursor.execute(f'''
                    UPDATE {vehicle_table}
                    SET Horometro = %s
                    WHERE VehicleID = (SELECT VehicleID FROM OrdenesdeSalida WHERE IdOrdenSalida = %s)
                ''', (horometroRegreso, order_id))

            conn.commit()
            flash('Orden de salida actualizada con éxito.', 'success')
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'Error al actualizar la orden de salida: {err}', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('departure_order_detail', company=company, order_id=order_id))

@app.route('/complete_departure_order/<company>/<int:order_id>', methods=['POST'])
def complete_departure_order(company, order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Mapear la tabla de vehículos según la empresa
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    horometro_regreso = request.form.get('horometroRegreso')

    try:
        # Obtener el VehicleID, VehicleName y HorometroSalida desde OrdenesdeSalida
        cursor.execute('''
            SELECT os.VehicleID, v.VehicleName, os.HorometroSalida
            FROM OrdenesdeSalida os
            JOIN {vehicle_table} v ON os.VehicleID = v.VehicleID
            WHERE os.IdOrdenSalida = %s AND os.Empresa = %s
        '''.format(vehicle_table=vehicle_table), (order_id, company))
        vehicle = cursor.fetchone()

        if vehicle:
            vehicle_id = vehicle['VehicleID']
            vehicle_name = vehicle['VehicleName']
            horometro_salida = vehicle['HorometroSalida']

            if vehicle_name.lower() != 'generico':
                if horometro_regreso is None or (horometro_salida is not None and float(horometro_regreso) < horometro_salida):
                    flash('El horómetro de regreso no puede ser menor que el horómetro de salida.', 'error')
                    return redirect(url_for('departure_order_detail', company=company, order_id=order_id))

                # Actualizar la orden de salida y el vehículo
                cursor.execute('''
                    UPDATE OrdenesdeSalida
                    SET HoraRegreso = NOW(), HorometroRegreso = %s
                    WHERE IdOrdenSalida = %s AND Empresa = %s
                ''', (horometro_regreso, order_id, company))

                new_horometro = float(horometro_regreso)

                cursor.execute(f'SELECT Horometro, HorometroDesdeUltimoMantenimiento FROM {vehicle_table} WHERE VehicleID = %s', (vehicle_id,))
                vehicle_data = cursor.fetchone()

                if vehicle_data:
                    current_horometro = vehicle_data['Horometro']
                    horometro_ultimo_mantenimiento = vehicle_data['HorometroDesdeUltimoMantenimiento']

                    if current_horometro is not None and new_horometro < current_horometro:
                        raise ValueError("El nuevo horómetro no puede ser menor que el valor actual.")
                    else:
                        current_date = datetime.now().strftime('%Y-%m-%d')
                        new_hdum = (new_horometro - current_horometro) if current_horometro is not None else new_horometro
                        new_hdum = horometro_ultimo_mantenimiento + new_hdum if horometro_ultimo_mantenimiento is not None else new_hdum

                        cursor.execute(f'''
                            UPDATE {vehicle_table}
                            SET 
                                Horometro = %s,
                                HorometroDesdeUltimoMantenimiento = %s,
                                FechaActualizacionHorometro = %s,
                                Status = CASE
                                            WHEN Status = 'Reparación Externa' THEN 'En Taller'
                                            WHEN Status = 'En Renta' THEN 'Disponible'
                                            ELSE Status
                                        END,
                                Ubicacion = %s
                            WHERE VehicleID = %s
                        ''', (new_horometro, new_hdum, current_date, company, vehicle_id))

            else:
                # Actualizar la orden de salida sin cambiar el vehículo genérico
                cursor.execute('''
                    UPDATE OrdenesdeSalida
                    SET HoraRegreso = NOW()
                    WHERE IdOrdenSalida = %s AND Empresa = %s
                ''', (order_id, company))

            conn.commit()
            flash('Orden de salida completada y vehículo actualizado.', 'success')
        else:
            flash('No se encontró el vehículo para la orden de salida especificada.', 'error')

    except (mysql.connector.Error, ValueError) as err:
        conn.rollback()
        flash(f'Error al completar la orden de salida: {err}', 'error')
        return redirect(url_for('departure_order_detail', company=company, order_id=order_id))

    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('active_departure_orders', company=company))





@app.route('/departure_orders_record/<company>', methods=['GET', 'POST'])
def departure_orders_record(company):
    order_id_query = request.args.get('order_id', '')
    operator_name_query = request.args.get('operator_name', '')
    vehicle_name_query = request.args.get('vehicle_name', '')
    client_name_query = request.args.get('client_name', '')
    start_date_query = request.args.get('start_date', '')
    end_date_query = request.args.get('end_date', '')
    page = int(request.args.get('page', 1))
    per_page = 20

    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        vehicle_table_map = {
            'MontasaHN': 'Vehicles',
            'MontasaCR': 'vehiculosCR',
            'Monhaco': 'vehiculosMonhaco'
        }
        vehicle_table = vehicle_table_map.get(company, 'Vehicles')

        cursor.execute('SELECT DISTINCT IdOrdenSalida FROM OrdenesdeSalida WHERE Empresa = %s ORDER BY IdOrdenSalida DESC', (company,))
        order_ids = cursor.fetchall()

        cursor.execute(f'SELECT DISTINCT VehicleName FROM {vehicle_table} WHERE Empresa = %s ORDER BY VehicleID', (company,))
        vehicle_names = cursor.fetchall()

        cursor.execute('SELECT DISTINCT NombreOperador FROM Operadores')
        operator_names = [row[0] for row in cursor.fetchall()]

        cursor.execute('SELECT DISTINCT ClienteID, Nombre, CodigoSap FROM Clientes')
        client_names = cursor.fetchall()

        query_conditions = ["os.HoraRegreso IS NOT NULL", "os.Empresa = %s"]
        query_params = [company]

        if order_id_query:
            query_conditions.append("os.IdOrdenSalida = %s")
            query_params.append(order_id_query)

        if operator_name_query:
            query_conditions.append("op.NombreOperador LIKE %s")
            query_params.append(f'%{operator_name_query}%')

        if vehicle_name_query:
            query_conditions.append("v.VehicleName LIKE %s")
            query_params.append(f'%{vehicle_name_query}%')

        if client_name_query:
            query_conditions.append("c.Nombre LIKE %s")
            query_params.append(f'%{client_name_query}%')

        if start_date_query:
            query_conditions.append("os.HoraCreado >= %s")
            query_params.append(start_date_query)

        if end_date_query:
            query_conditions.append("os.HoraCreado <= %s")
            query_params.append(end_date_query)

        where_clause = " AND ".join(query_conditions)
        count_query = f'''
            SELECT COUNT(*) as total
            FROM OrdenesdeSalida os
            JOIN {vehicle_table} v ON os.VehicleID = v.VehicleID
            JOIN Operadores op ON os.IdOperador = op.IdOperador
            JOIN Clientes c ON os.ClienteID = c.ClienteID
            WHERE {where_clause}
        '''
        cursor.execute(count_query, tuple(query_params))
        total_results = cursor.fetchone()[0]

        total_pages = (total_results // per_page) + (1 if total_results % per_page > 0 else 0)
        offset = (page - 1) * per_page

        sql_query = f'''
            SELECT os.IdOrdenSalida, os.VehicleID, v.VehicleName, os.IdOperador, op.NombreOperador,
                   os.HoraCreado, os.Comentarios, os.Ubicacion, os.HoraRegreso, os.OrdenDeSAP, c.Nombre as ClienteNombre
            FROM OrdenesdeSalida os
            JOIN {vehicle_table} v ON os.VehicleID = v.VehicleID
            JOIN Operadores op ON os.IdOperador = op.IdOperador
            JOIN Clientes c ON os.ClienteID = c.ClienteID
            WHERE {where_clause}
            ORDER BY os.IdOrdenSalida DESC
            LIMIT {per_page} OFFSET {offset}
        '''
        cursor.execute(sql_query, tuple(query_params))
        departure_orders = cursor.fetchall()

    except mysql.connector.Error as err:
        print("SQL Error: ", err)
        departure_orders = []
    finally:
        cursor.close()
        conn.close()

    return render_template('departure_orders_record.html',
                           departure_orders=departure_orders,
                           order_ids=order_ids,
                           vehicle_names=vehicle_names,
                           operator_names=operator_names,
                           client_names=client_names,
                           company=company,
                           page=page,
                           total_pages=total_pages,
                           selected_order_id=order_id_query,
                           selected_vehicle_name=vehicle_name_query,
                           selected_operator_name=operator_name_query,
                           selected_client_name=client_name_query,
                           start_date=start_date_query,
                           end_date=end_date_query)


       



@app.route('/departure_order_record_detail/<company>/<int:order_id>')
def departure_order_record_detail(company, order_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Mapeo de la compañía a la tabla de vehículos correspondiente
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')  # Usa 'Vehicles' como predeterminado

    # Consultar los detalles de la orden de salida específica utilizando la tabla de vehículos correcta
    cursor.execute(f'''
        SELECT os.IdOrdenSalida, os.VehicleID, v.VehicleName, os.IdOperador, op.NombreOperador,
               os.HoraCreado, os.Comentarios, os.Ubicacion, os.HorometroSalida, os.HorometroRegreso, os.HoraRegreso, os.OrdenDeSAP, c.Nombre as ClienteNombre
        FROM OrdenesdeSalida os
        JOIN {vehicle_table} v ON os.VehicleID = v.VehicleID
        JOIN Operadores op ON os.IdOperador = op.IdOperador
        JOIN Clientes c ON os.ClienteID = c.ClienteID
        WHERE os.IdOrdenSalida = %s AND v.Empresa = %s
    ''', (order_id, company))
    order_detail = cursor.fetchone()

    cursor.close()
    conn.close()

    if order_detail:
        # Pasar los detalles de la orden a la plantilla HTML
        return render_template('departure_order_record_detail.html', order=order_detail, company=company, order_id=order_id)
    else:
        # En caso de que la orden no exista, mostrar un mensaje o redirigir
        flash('Orden de salida no encontrada.', 'error')
        return redirect(url_for('departure_orders_record', company=company))











#Ordenes de SALIDA Fin

#horometros
def update_horometros_via_email(vehicle_name, new_horometro):
    conn = get_db_connection()
    cursor = conn.cursor()
    error_message = None
    company = 'MontasaHN'  # Asume que siempre será MontasaHN en este caso, ajusta según sea necesario.

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Obtener el Vehicle ID desde el nombre del vehículo
    cursor.execute(f'SELECT VehicleID FROM {vehicle_table} WHERE VehicleName = %s', (vehicle_name,))
    result = cursor.fetchone()

    if result is None:
        return {"error": "Vehicle data not found."}

    vehicle_id = result[0]

    if new_horometro is not None:
        new_horometro = float(new_horometro)

        # Obtener el horómetro actual para el vehículo
        cursor.execute(f'SELECT Horometro, HorometroDesdeUltimoMantenimiento FROM {vehicle_table} WHERE VehicleID = %s', (vehicle_id,))
        vehicle_data = cursor.fetchone()

        if vehicle_data:
            current_horometro = vehicle_data[0]
            horometro_ultimo_mantenimiento = vehicle_data[1]

            if current_horometro is not None and new_horometro < current_horometro:
                return {"error": "El nuevo horómetro no puede ser menor que el valor actual."}
            else:
                current_date = datetime.now().strftime('%Y-%m-%d')
                # Calcular el nuevo HorometroDesdeUltimoMantenimiento correctamente
                new_hdum = (new_horometro - current_horometro) if current_horometro is not None else new_horometro
                new_hdum = horometro_ultimo_mantenimiento + new_hdum if horometro_ultimo_mantenimiento is not None else new_hdum

                # Actualizar el horómetro y la fecha de actualización
                cursor.execute(f'UPDATE {vehicle_table} SET Horometro = %s, HorometroDesdeUltimoMantenimiento = %s, FechaActualizacionHorometro = %s WHERE VehicleID = %s',
                            (new_horometro, new_hdum, current_date, vehicle_id))
                conn.commit()

                
                send_horometro_update(vehicle_name, new_horometro)

        else:
            return {"error": "Vehicle data not found."}

    cursor.close()
    conn.close()


@app.route('/update_horometros', methods=['POST'])
def update_horometros():
    conn = get_db_connection()
    cursor = conn.cursor()
    error_message = None

    vehicle_name = request.form.get('vehicle_name')
    new_horometro = request.form.get('new_horometro')
    company = request.form.get('company')

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    # Get Vehicle ID from Vehicle Name
    cursor.execute(f'SELECT VehicleID FROM {vehicle_table} WHERE VehicleName = %s', (vehicle_name,))
    result = cursor.fetchone()

    if result is None:
        referrer = request.referrer or url_for('vehicles_list', company=company)
        return redirect(f"{referrer}?error_message=Vehicle data not found.")
    
    vehicle_id = result[0]

    if new_horometro is not None:
        new_horometro = float(new_horometro)

        # Fetch current horometro for the vehicle
        cursor.execute(f'SELECT Horometro, HorometroDesdeUltimoMantenimiento FROM {vehicle_table} WHERE VehicleID = %s', (vehicle_id,))
        vehicle_data = cursor.fetchone()

        if vehicle_data:
            current_horometro = vehicle_data[0]
            horometro_ultimo_mantenimiento = vehicle_data[1]

            if current_horometro is not None and new_horometro < current_horometro:
                error_message = "El nuevo horómetro no puede ser menor que el valor actual."
            else:
                current_date = datetime.now().strftime('%Y-%m-%d')
                # Calculate HorometroDesdeUltimoMantenimiento correctly
                new_hdum = (new_horometro - current_horometro) if current_horometro is not None else new_horometro
                new_hdum = horometro_ultimo_mantenimiento + new_hdum if horometro_ultimo_mantenimiento is not None else new_hdum

                # Update both Horometro and FechaActualizacionHorometro
                cursor.execute(f'UPDATE {vehicle_table} SET Horometro = %s, HorometroDesdeUltimoMantenimiento = %s, FechaActualizacionHorometro = %s WHERE VehicleID = %s',
                            (new_horometro, new_hdum, current_date, vehicle_id))
                conn.commit()

                if company == 'MontasaHN':
                    send_horometro_update(vehicle_name, new_horometro)

        else:
            referrer = request.referrer or url_for('vehicles_list', company=company)
            return redirect(f"{referrer}?error_message=Vehicle data not found.")

    cursor.close()
    conn.close()

    # Redirect back to the previous page with error message if any
    referrer = request.referrer or url_for('vehicles_list', company=company)
    if error_message:
        flash(error_message, 'error')
    return redirect(referrer)







@app.route('/update_vehicle/<int:vehicle_id>', methods=['POST'])
def update_vehicle_locacion_disponibilidad(vehicle_id):
    Ubicacion = request.form['Ubicacion']
    Disponibilidad = request.form['Disponibilidad']
    company = request.form['company']  # Capture the company from the form

    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Determine the appropriate vehicle table
        vehicle_table_map = {
            'MontasaHN': 'Vehicles',
            'MontasaCR': 'vehiculosCR',
            'Monhaco': 'vehiculosMonhaco'
        }
        vehicle_table = vehicle_table_map.get(company, 'Vehicles')

        update_query = f'''
            UPDATE {vehicle_table}
            SET Ubicacion = %s, Disponibilidad = %s
            WHERE VehicleID = %s
        '''
        cursor.execute(update_query, (Ubicacion, Disponibilidad, vehicle_id))
        conn.commit()
        flash('Vehicle updated successfully.', 'success')
    except mysql.connector.Error as err:
        conn.rollback()
        flash(f'Error updating vehicle: {err.msg}', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('vehicles_list', company=company))


@app.route('/vehicle_search')
def vehicle_search():
    search_term = request.args.get('q', '')  # 'q' es un nombre común para parámetros de búsqueda
    company = request.args.get('company', 'MontasaHN')  # Default to MontasaHN if no company is specified
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(f"SELECT VehicleID, VehicleName FROM {vehicle_table} WHERE VehicleName LIKE %s", ('%' + search_term + '%',))
    vehicles = cursor.fetchall()
    cursor.close()
    conn.close()
    
    results = [{'id': vehicle[0], 'text': vehicle[1]} for vehicle in vehicles]  # Asegúrate de que esto coincide con tu estructura de datos
    return jsonify(results)




#horometros
#reporte ubicacion
import gc

# def generate_vehicle_report(company):
#     original_file_path = 'C:\\Users\\rodrigo.monterroso\\Downloads\\flotas first running 7.9.2024\\flotas Upload 30 6 24\\Flotas\\15-07-2024 REPORTE UBICACION MONTASA -.xlsx'
#     copied_file_path = 'C:\\Users\\rodrigo.monterroso\\Downloads\\flotas first running 7.9.2024\\flotas Upload 30 6 24\\Flotas\\temp_REPORTE_UBICACION_MONTASA.xlsx'
    
#     # Copy the original file to a new location
#     #logging.info("Creating a writable copy of the Excel file...")
#     shutil.copyfile(original_file_path, copied_file_path)

#     conn = get_db_connection()
#     cursor = conn.cursor()

#     vehicle_table_map = {
#         'MontasaHN': 'Vehicles',
#         'MontasaCR': 'vehiculosCR',
#         'Monhaco': 'vehiculosMonhaco'
#     }
#     vehicle_table = vehicle_table_map.get(company, 'Vehicles')

#     # Fetch vehicle data
#     #logging.info("Fetching vehicle data from database...")
#     cursor.execute(f'''
#         SELECT VehicleName, Observacion, Status, Ubicacion
#         FROM {vehicle_table}
#         WHERE VehicleName NOT LIKE 'C%' AND VehicleName != 'generico'
#     ''')
#     vehicles = cursor.fetchall()
    

#     try:
#         # Load existing workbook in read-only mode to create mapping
#         #logging.info("Loading workbook...")
#         workbook = load_workbook(copied_file_path, data_only=True, read_only=True)
#         sheet = workbook.active
#         #logging.info("Workbook loaded successfully.")
#         #logging.info("Active sheet selected.")

#         # Create a mapping of vehicle names to their respective rows
#         #logging.info("Creating vehicle row mapping...")
#         vehicle_row_mapping = {}
#         for row in sheet.iter_rows(min_row=8, max_row=196, max_col=26):
#             vehicle_name = row[7].value  # Column H (8th index) is 'No. EQUIPO'
#             if vehicle_name:
#                 vehicle_row_mapping[vehicle_name] = row
#         #logging.info("Vehicle row mapping created.")

#         # Close the read-only workbook
#         workbook.close()

#         # Load the workbook again in write mode
#         #logging.info("Loading workbook in write mode...")
#         workbook = load_workbook(copied_file_path)
#         sheet = workbook.active

#         # Update each vehicle row by row
#         #logging.info("Updating vehicle data row by row...")
#         red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
#         yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
#         for vehicle in vehicles:
#             vehicle_name = vehicle[0]
#             if vehicle_name in vehicle_row_mapping:
#                 row = vehicle_row_mapping[vehicle_name]
#                 try:
#                     sheet.cell(row=row[0].row, column=14).value = '1' if vehicle[2] in ['En Taller', 'Deshabilitado'] else '2' if vehicle[2] in ['En Renta', 'Reparacion Externa'] else '3' if vehicle[2] == 'Disponible' else ''  # ST
#                     #logging.info(f"Updated ST for {vehicle_name}")
#                 except Exception as e:
#                     logging.error(f"Error updating ST for {vehicle_name}: {e}")
#                 try:
#                     sheet.cell(row=row[0].row, column=15).value = vehicle[1]  # OBSERVACIONES
#                     #logging.info(f"Updated OBSERVACIONES for {vehicle_name}")
#                 except Exception as e:
#                     logging.error(f"Error updating OBSERVACIONES for {vehicle_name}: {e}")
#                 try:
#                     sheet.cell(row=row[0].row, column=18).value = vehicle[3]  # UBICACIÓN
#                     #logging.info(f"Updated UBICACIÓN for {vehicle_name}")
#                 except Exception as e:
#                     logging.error(f"Error updating UBICACIÓN for {vehicle_name}: {e}")

#                 # Highlight rows and update 'deshabilitado' column
#                 try:
#                     if vehicle[2] == 'Deshabilitado':
#                         for cell in row:
#                             sheet.cell(row=row[0].row, column=cell.col_idx).fill = red_fill
#                         sheet.cell(row=row[0].row, column=23).value = 1  # deshabilitado column (column W is the 23rd column, index 22)
#                         #logging.info(f"Row for {vehicle_name} highlighted in red for being disabled.")
#                     elif 'VENDIDO' in vehicle[1]:
#                         for cell in row:
#                             sheet.cell(row=row[0].row, column=cell.col_idx).fill = yellow_fill
#                         #logging.info(f"Row for {vehicle_name} highlighted in yellow for being sold.")
#                 except Exception as e:
#                     logging.error(f"Error updating highlight for {vehicle_name}: {e}")

#         # Explicitly call garbage collection
#         gc.collect()

#         # Save the workbook to a new file with the current date

#         current_date = datetime.now().strftime('%d-%m-%Y')
#         new_file_path = os.path.join(os.path.dirname(copied_file_path), f'{current_date} REPORTE UBICACION MONTASA -.xlsx')
#         workbook.save(new_file_path)
#         #logging.info(f"Workbook saved to {new_file_path}")

#         # Close the cursor and connection
#         #logging.info("Closing database connection...")
#         cursor.close()
#         conn.close()

#         return new_file_path

#     except Exception as e:
#         logging.error(f"Error processing the request: {e}")
#         return None

@app.route('/download_vehicle_report/<company>', methods=['GET'])
def download_vehicle_report(company):
    file_path = generate_vehicle_report(company)
    if file_path:
        return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))
    else:
        return "Internal Server Error", 500

# Suppress OpenPyXL warnings
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")




@app.route('/download_active_order/<int:order_id>', methods=['GET'])
def download_active_order(order_id):
    # Fetch order details
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT wo.OrderID, wo.VehicleID, IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName, wo.WorkType, wo.Description, wo.Status,
            wo.Lugar, IFNULL(c.Nombre, wo.Dueno) AS Dueno, wo.Marca, wo.Diagnostico, wo.TrabajoRealizado, wo.CreatedTime, wo.Empresa, wo.currently_waiting
        FROM WorkOrders wo
        LEFT JOIN Vehicles v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        LEFT JOIN Clientes c ON wo.ClienteCodigoSap = c.CodigoSap
        WHERE wo.OrderID = %s
    ''', (order_id,))
    order = cursor.fetchone()

    if not order:
        cursor.close()
        conn.close()
        return 'Work Order not found', 404

    company = order[12]
    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    cursor.execute(f'SELECT Horometro FROM {vehicle_table} WHERE VehicleID = %s', (order[1],))
    vehicle_row = cursor.fetchone()
    horometro = vehicle_row[0] if vehicle_row else 0

    cursor.execute('''
        SELECT m.Name
        FROM Mechanics m
        JOIN MechanicWorkOrder mwo ON m.MechanicID = mwo.MechanicID
        WHERE mwo.OrderID = %s
    ''', (order_id,))
    assigned_mechanics = [row[0] for row in cursor.fetchall()]

    cursor.execute('''
        SELECT CodigoSap, Description, Quantity
        FROM WorkOrderSupplies
        WHERE OrderID = %s AND Status = 'Listo'
    ''', (order_id,))
    assigned_supplies = cursor.fetchall()

    cursor.close()
    conn.close()

    data = {
        'Lugar': order[6],
        'FechaHoraRecibido': order[11].strftime('%Y-%m-%d %H:%M:%S'),
        'NumeroVehiculo': order[2],
        'OrdenTrabajoId': order[0],
        'Marca': order[8],
        'Modelo': '',  # Fill this if you have this data
        'Serie': '',  # Fill this if you have this data
        'Horometro': horometro,
        'Descripcion': order[4] if order[4] is not None else '',  # Use Description field from the work order
        'Diagnostico': order[9] if order[9] is not None else '',
        'TrabajoDesempenado': order[10] if order[10] is not None else '',
        'Mecanicos': ', '.join(assigned_mechanics),
        'RepuestosAnadidos': [{'Cantidad': repuesto[2], 'Codigo': repuesto[0], 'Descripcion': repuesto[1]} for repuesto in assigned_supplies]
    }

    template_path = r'C:\Users\rodrigo.monterroso\Downloads\flotas first running 7.9.2024\flotas Upload 30 6 24\Flotas\FORMATO ORDEN DE SERVICIO-GM-FO-0104.XLSX'
    output_path = os.path.join(os.path.dirname(template_path), f'Orden_{order[0]}_{order[2]}.xlsx')

    # Load the workbook and select the active sheet
    wb = load_workbook(template_path)
    sheet = wb.active

    # Fill in the data
    sheet['C8'] = data['Lugar']
    sheet['O8'] = data['FechaHoraRecibido']
    sheet['D9'] = data['NumeroVehiculo']
    sheet['X9'] = data['OrdenTrabajoId']
    sheet['C11'] = data['Marca']
    sheet['I11'] = data['Modelo']
    sheet['P11'] = data['Serie']
    sheet['X11'] = data['Horometro']
    sheet['G14'] = data['Descripcion']

    def split_text(text, max_chars):
        words = text.split(' ')
        lines = []
        current_line = []
        current_length = 0

        for word in words:
            if current_length + len(word) + 1 > max_chars:
                lines.append(' '.join(current_line))
                current_line = [word]
                current_length = len(word) + 1
            else:
                current_line.append(word)
                current_length += len(word) + 1

        if current_line:
            lines.append(' '.join(current_line))
        return lines

    diagnostico_lines = split_text(data['Diagnostico'], 60)
    trabajo_lines = split_text(data['TrabajoDesempenado'], 60)

    for i, line in enumerate(diagnostico_lines[:9]):
        sheet[f'A{31 + i}'] = line
    for i, line in enumerate(trabajo_lines[:7]):
        sheet[f'A{41 + i}'] = line

    mecanico_cells = ['A49', 'H49', 'O49', 'V49', 'A50', 'H50', 'O50', 'V50', 'A51', 'H51', 'O51', 'V51']
    for i, mecanico in enumerate(data['Mecanicos'].split(', ')[:12]):
        cell = mecanico_cells[i]
        sheet[cell] = mecanico
        sheet[cell].font = Font(size=20)

    for i, repuesto in enumerate(data['RepuestosAnadidos'][:14]):
        sheet[f'B{55 + i}'] = repuesto['Cantidad']
        sheet[f'D{55 + i}'] = repuesto['Codigo']
        sheet[f'H{55 + i}'] = repuesto['Descripcion']

    # Save the filled workbook
    wb.save(output_path)

    # Send the file as a download
    response = send_file(output_path, as_attachment=True, download_name=f'Orden_{order[0]}_{order[2]}.xlsx')

    # Delete the created files after sending
    @response.call_on_close
    def cleanup():
        os.remove(output_path)

    return response

@app.route('/download_record_order/<company>/<int:order_id>', methods=['GET'])
def download_record_order(company, order_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    vehicle_table_map = {
        'MontasaHN': 'Vehicles',
        'MontasaCR': 'vehiculosCR',
        'Monhaco': 'vehiculosMonhaco'
    }
    vehicle_table = vehicle_table_map.get(company, 'Vehicles')

    cursor.execute(f'''
        SELECT wo.OrderID, wo.VehicleID, IFNULL(v.VehicleName, vc.VehicleName) AS VehicleName, wo.WorkType,
               wo.Description, wo.CreatedTime, wo.FinishedTime, wo.Lugar, wo.Dueno, wo.Marca,
               wo.Diagnostico, wo.TrabajoRealizado, wo.WorkedTime
        FROM WorkOrders wo
        LEFT JOIN {vehicle_table} v ON wo.VehicleID = v.VehicleID AND wo.ClienteCodigoSap IS NULL
        LEFT JOIN VehiclesClientes vc ON wo.VehicleID = vc.VehicleID AND wo.ClienteCodigoSap IS NOT NULL
        WHERE wo.OrderID = %s AND wo.Status = 'Completed' AND wo.Empresa = %s
    ''', (order_id, company))
    order = cursor.fetchone()

    if not order:
        cursor.close()
        conn.close()
        return 'Work Order not found', 404

    cursor.execute(f'SELECT Horometro FROM {vehicle_table} WHERE VehicleID = %s', (order[1],))
    vehicle_row = cursor.fetchone()
    horometro = vehicle_row[0] if vehicle_row else 0

    cursor.execute('''
        SELECT m.Name
        FROM Mechanics m
        JOIN TimeTracking t ON m.MechanicID = t.MechanicID
        WHERE t.OrderID = %s
    ''', (order_id,))
    assigned_mechanics = [row[0] for row in cursor.fetchall()]

    cursor.execute('''
        SELECT s.CodigoSap, s.Description, ws.Quantity
        FROM WorkOrderSupplies ws
        JOIN Supplies s ON ws.CodigoSap = s.CodigoSap
        WHERE ws.OrderID = %s AND ws.Status = 'Recibido'
    ''', (order_id,))
    assigned_supplies = cursor.fetchall()

    cursor.close()
    conn.close()

    data = {
        'Lugar': order[7],
        'FechaHoraRecibido': order[5].strftime('%Y-%m-%d %H:%M:%S'),
        'NumeroVehiculo': order[2],
        'OrdenTrabajoId': order[0],
        'Marca': order[9],
        'Modelo': '',  # Fill this if you have this data
        'Serie': '',  # Fill this if you have this data
        'Horometro': horometro,
        'Descripcion': order[4] if order[4] is not None else '',  # Use Description field from the work order
        'Diagnostico': order[10] if order[10] is not None else '',
        'TrabajoDesempenado': order[11] if order[11] is not None else '',
        'Mecanicos': ', '.join(assigned_mechanics),
        'RepuestosAnadidos': [{'CodigoSap': repuesto[0], 'Cantidad': repuesto[2], 'Descripcion': repuesto[1]} for repuesto in assigned_supplies]
    }

    template_path = r'C:\Users\rodrigo.monterroso\Downloads\flotas first running 7.9.2024\flotas Upload 30 6 24\Flotas\FORMATO ORDEN DE SERVICIO-GM-FO-0104.XLSX'
    output_path = os.path.join(os.path.dirname(template_path), f'Orden_{order[0]}_{order[2]}.xlsx')

    wb = load_workbook(template_path)
    sheet = wb.active

    sheet['C8'] = data['Lugar']
    sheet['O8'] = data['FechaHoraRecibido']
    sheet['D9'] = data['NumeroVehiculo']
    sheet['X9'] = data['OrdenTrabajoId']
    sheet['C11'] = data['Marca']
    sheet['I11'] = data['Modelo']
    sheet['P11'] = data['Serie']
    sheet['X11'] = data['Horometro']
    sheet['G14'] = data['Descripcion']

    def split_text(text, max_chars):
        words = text.split(' ')
        lines = []
        current_line = []
        current_length = 0

        for word in words:
            if current_length + len(word) + 1 > max_chars:
                lines.append(' '.join(current_line))
                current_line = [word]
                current_length = len(word) + 1
            else:
                current_line.append(word)
                current_length += len(word) + 1

        if current_line:
            lines.append(' '.join(current_line))
        return lines

    diagnostico_lines = split_text(data['Diagnostico'], 60)
    trabajo_lines = split_text(data['TrabajoDesempenado'], 60)

    for i, line in enumerate(diagnostico_lines[:9]):
        sheet[f'A{31 + i}'] = line
    for i, line in enumerate(trabajo_lines[:7]):
        sheet[f'A{41 + i}'] = line

    mecanico_cells = ['A49', 'H49', 'O49', 'V49', 'A50', 'H50', 'O50', 'V50', 'A51', 'H51', 'O51', 'V51']
    for i, mecanico in enumerate(data['Mecanicos'].split(', ')[:12]):
        cell = mecanico_cells[i]
        sheet[cell] = mecanico
        sheet[cell].font = Font(size=20)

    for i, repuesto in enumerate(data['RepuestosAnadidos'][:14]):
        sheet[f'B{55 + i}'] = repuesto['Cantidad']
        sheet[f'D{55 + i}'] = repuesto['CodigoSap']
        sheet[f'H{55 + i}'] = repuesto['Descripcion']

    wb.save(output_path)

    response = send_file(output_path, as_attachment=True, download_name=f'Orden_{order[0]}_{order[2]}.xlsx')

    @response.call_on_close
    def cleanup():
        os.remove(output_path)

    return response


















#excels
#pin

@app.route('/validate_mechanic_pin')
def validate_mechanic_pin():
    pin = request.args.get('pin')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT MechanicID, Name FROM Mechanics WHERE PinCode = %s', (pin,))
    mechanic = cursor.fetchone()
    cursor.close()
    conn.close()
    if mechanic:
        return jsonify({'id': mechanic['MechanicID'], 'name': mechanic['Name']})
    else:
        return jsonify({'id': None, 'name': None})

@app.route('/validate_operator_pin')
def validate_operator_pin():
    pin = request.args.get('pin')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT IdOperador, NombreOperador FROM Operadores WHERE PinCode = %s', (pin,))
    operator = cursor.fetchone()
    cursor.close()
    conn.close()
    if operator:
        return jsonify({'id': operator['IdOperador'], 'name': operator['NombreOperador']})
    else:
        return jsonify({'id': None, 'name': None})

@app.route('/validate_boss_pin')
def validate_boss_pin():
    pin = request.args.get('pin')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT BossID, Nombre FROM Boss WHERE BossPin = %s', (pin,))
    boss = cursor.fetchone()
    cursor.close()
    conn.close()
    if boss:
        return jsonify({'id': boss['BossID'], 'name': boss['Nombre']})
    else:
        return jsonify({'id': None, 'name': None})



@app.route('/get_mechanic_name')
def get_mechanic_name():
    mechanic_id = request.args.get('id')
    if mechanic_id is None:
        return jsonify({'name': None})

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT Name FROM Mechanics WHERE MechanicID = %s', (mechanic_id,))
    mechanic = cursor.fetchone()
    cursor.close()
    conn.close()

    if mechanic:
        return jsonify({'name': mechanic['Name']})
    else:
        return jsonify({'name': None})

@app.route('/get_boss_name')
def get_boss_name():
    boss_id = request.args.get('id')
    if boss_id is None:
        return jsonify({'name': None})

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT Nombre FROM Boss WHERE BossID = %s', (boss_id,))
    boss = cursor.fetchone()
    cursor.close()
    conn.close()

    if boss:
        return jsonify({'name': boss['Nombre']})
    else:
        return jsonify({'name': None})

#pin


#excel






##en el server correr asi :
from waitress import serve
from app import app

if __name__ == '__main__':
    with app.app_context():
        initialize_scheduler()
        serve(app, host='0.0.0.0', port=8080)


# if __name__ == '__main__': #en la compu
#     app.run(debug=True)

# .\venv\Scripts\Activate
# flask run --debug - en la compu
# python app.py  - en el server


